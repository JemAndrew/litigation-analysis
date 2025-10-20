#!/usr/bin/env python3
"""
PHL Document Ingestion Script - Phase 1

Scans PHL disclosure folders, extracts document metadata (date and description),
and appends to existing Excel tracking sheet after C-156.

Features:
- Recursive folder scanning for all PHL documents
- PDF metadata extraction (creation date, description)
- Duplicate detection and flagging
- Excel formatting preservation using openpyxl
- Comprehensive error handling and logging
- Progress indicators for long operations

British English throughout.
"""

import os
import re
import logging
from pathlib import Path
from typing import List, Dict, Optional, Tuple
from datetime import datetime
from collections import defaultdict

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import PyPDF2
from tqdm import tqdm

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('phl_ingestion.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)


class PHLDocumentIngester:
    """
    Ingests PHL documents from disclosure folders into Excel tracking sheet.
    
    Preserves all Excel formatting while adding new document rows.
    """
    
    def __init__(
        self,
        excel_path: str,
        output_folder: str,
        base_phl_path: str
    ):
        """
        Initialise PHL document ingester.
        
        Args:
            excel_path: Path to source Excel file
            output_folder: Path to output directory
            base_phl_path: Base path for PHL disclosure folders
        """
        self.excel_path = Path(excel_path)
        self.output_folder = Path(output_folder)
        self.base_phl_path = Path(base_phl_path)
        
        self.phl_documents = []
        self.duplicates = []
        self.corrupted_files = []
        self.missing_metadata = []
        
        # Ensure output folder exists
        self.output_folder.mkdir(parents=True, exist_ok=True)
        
        logger.info("=" * 80)
        logger.info("PHL DOCUMENT INGESTION - PHASE 1")
        logger.info("=" * 80)
        logger.info(f"Excel source: {self.excel_path}")
        logger.info(f"Output folder: {self.output_folder}")
        logger.info(f"PHL base path: {self.base_phl_path}")
    
    def extract_phl_number(self, filename: str) -> Optional[int]:
        """
        Extract PHL number from filename.
        
        Args:
            filename: PDF filename (e.g., 'PHL_000123.pdf')
        
        Returns:
            PHL number as integer, or None if invalid
        """
        match = re.match(r'PHL_(\d+)\.pdf$', filename, re.IGNORECASE)
        if match:
            return int(match.group(1))
        return None
    
    def should_exclude_file(self, filename: str) -> bool:
        """
        Check if file should be excluded based on naming patterns.
        
        Args:
            filename: PDF filename
        
        Returns:
            True if file should be excluded
        """
        filename_lower = filename.lower()
        exclusions = ['redacted', 'v2', '_v2', ' v2']
        
        for exclusion in exclusions:
            if exclusion in filename_lower:
                logger.info(f"   Excluding {filename} (contains '{exclusion}')")
                return True
        
        return False
    
    def determine_production_date(self, file_path: Path) -> str:
        """
        Determine 'When Produced?' date based on folder location.
        
        Args:
            file_path: Full path to PHL document
        
        Returns:
            Production date string (DD Month YYYY format)
        """
        path_str = str(file_path)
        
        # Check for April 2025 production folder
        if '2025.04.11 PHL Production' in path_str:
            return '11 April 2025'
        
        # Check for May 2025 (or other folders in main disclosure)
        if '02. PHL Disclosure prior to 23 June' in path_str:
            return '28 May 2025'
        
        # Unknown location
        return ''
    
    def extract_pdf_creation_date(self, pdf_path: Path) -> Optional[str]:
        """
        Extract creation date from PDF metadata.
        
        Args:
            pdf_path: Path to PDF file
        
        Returns:
            Formatted date string (DD Month YYYY), or None if extraction fails
        """
        try:
            with open(pdf_path, 'rb') as f:
                pdf_reader = PyPDF2.PdfReader(f)
                metadata = pdf_reader.metadata
                
                if not metadata:
                    return None
                
                # Try creation date first
                date_field = metadata.get('/CreationDate') or metadata.get('/ModDate')
                
                if not date_field:
                    return None
                
                # Parse PDF date format: D:YYYYMMDDHHmmSS...
                # Example: D:20240328154523+00'00'
                date_match = re.match(r'D:(\d{4})(\d{2})(\d{2})', str(date_field))
                
                if date_match:
                    year, month, day = date_match.groups()
                    date_obj = datetime(int(year), int(month), int(day))
                    
                    # Format as DD Month YYYY
                    return date_obj.strftime('%d %B %Y')
                
                return None
                
        except Exception as e:
            logger.warning(f"   Could not extract date from {pdf_path.name}: {e}")
            return None
    
    def extract_pdf_description(self, pdf_path: Path) -> Optional[str]:
        """
        Extract description from PDF metadata or content.
        
        Tries multiple strategies:
        1. PDF title metadata
        2. PDF subject metadata
        3. First page text (first 200 characters)
        
        Args:
            pdf_path: Path to PDF file
        
        Returns:
            Description string, or None if extraction fails
        """
        try:
            with open(pdf_path, 'rb') as f:
                pdf_reader = PyPDF2.PdfReader(f)
                metadata = pdf_reader.metadata
                
                # Strategy 1: Check title metadata
                if metadata and metadata.get('/Title'):
                    title = str(metadata['/Title']).strip()
                    if title and title.lower() != 'untitled' and len(title) > 3:
                        return title[:200]  # Limit length
                
                # Strategy 2: Check subject metadata
                if metadata and metadata.get('/Subject'):
                    subject = str(metadata['/Subject']).strip()
                    if subject and len(subject) > 3:
                        return subject[:200]
                
                # Strategy 3: Extract from first page content
                if len(pdf_reader.pages) > 0:
                    first_page = pdf_reader.pages[0]
                    text = first_page.extract_text()
                    
                    if text:
                        # Clean and truncate
                        text = ' '.join(text.split())  # Normalise whitespace
                        text = text[:200]  # Limit length
                        
                        if len(text) > 20:  # Meaningful content
                            return text
                
                return None
                
        except Exception as e:
            logger.warning(f"   Could not extract description from {pdf_path.name}: {e}")
            return None
    
    def scan_phl_directories(self) -> List[Dict]:
        """
        Recursively scan PHL disclosure folders for all PHL documents.
        
        Returns:
            List of dictionaries with PHL document information
        """
        logger.info("\n" + "=" * 80)
        logger.info("SCANNING PHL DIRECTORIES")
        logger.info("=" * 80)
        
        # Main search path (recursive)
        search_path = self.base_phl_path / "02. PHL Disclosure prior to 23 June"
        
        if not search_path.exists():
            logger.error(f"ERROR: Path does not exist: {search_path}")
            raise FileNotFoundError(f"PHL disclosure path not found: {search_path}")
        
        logger.info(f"Searching recursively in: {search_path}")
        
        # Find all PHL PDFs
        phl_files = {}  # {phl_number: [list of paths]}
        
        for pdf_file in search_path.rglob('*.pdf'):
            # Check if it's a PHL file
            phl_number = self.extract_phl_number(pdf_file.name)
            
            if phl_number is None:
                continue
            
            # Check exclusions
            if self.should_exclude_file(pdf_file.name):
                continue
            
            # Store (may have duplicates)
            if phl_number not in phl_files:
                phl_files[phl_number] = []
            phl_files[phl_number].append(pdf_file)
        
        logger.info(f"Found {len(phl_files)} unique PHL numbers")
        logger.info(f"Total PHL files found: {sum(len(paths) for paths in phl_files.values())}")
        
        # Process documents
        logger.info("\n" + "=" * 80)
        logger.info("PROCESSING PHL DOCUMENTS")
        logger.info("=" * 80)
        
        phl_documents = []
        
        # Sort by PHL number
        for phl_number in tqdm(sorted(phl_files.keys()), desc="Processing PHL docs"):
            paths = phl_files[phl_number]
            
            # Use first occurrence
            primary_path = paths[0]
            
            # Flag duplicates
            notes = []
            if len(paths) > 1:
                self.duplicates.append({
                    'phl_number': phl_number,
                    'paths': paths
                })
                notes.append(f"DUPLICATE - Found in {len(paths)} locations")
                logger.info(f"   PHL_{phl_number:06d}: Duplicate detected ({len(paths)} copies)")
            
            # Extract metadata
            date = self.extract_pdf_creation_date(primary_path)
            description = self.extract_pdf_description(primary_path)
            
            # Handle extraction failures
            if date is None:
                date = "ERROR: Could not extract"
                self.missing_metadata.append(phl_number)
            
            if description is None:
                description = "ERROR: Could not extract"
                if phl_number not in self.missing_metadata:
                    self.missing_metadata.append(phl_number)
            
            # Determine production date
            when_produced = self.determine_production_date(primary_path)
            
            phl_documents.append({
                'phl_number': phl_number,
                'reference': f'PHL_{phl_number:06d}',
                'date': date,
                'description': description,
                'trial_bundle_ref': '',
                'trial_bundle_index_desc': '',
                'when_produced': when_produced,
                'notes': '; '.join(notes) if notes else '',
                'path': str(primary_path)
            })
        
        logger.info(f"\n✅ Processed {len(phl_documents)} PHL documents")
        logger.info(f"   Duplicates detected: {len(self.duplicates)}")
        logger.info(f"   Metadata extraction issues: {len(self.missing_metadata)}")
        
        return phl_documents
    
    def append_to_excel(self, phl_docs: List[Dict]) -> Path:
        """
        Append PHL documents to Excel, preserving all formatting.
        
        Uses openpyxl to maintain cell styles, borders, colours, etc.
        
        Args:
            phl_docs: List of PHL document dictionaries
        
        Returns:
            Path to output Excel file
        """
        logger.info("\n" + "=" * 80)
        logger.info("UPDATING EXCEL FILE")
        logger.info("=" * 80)
        
        # Load workbook
        logger.info(f"Loading Excel: {self.excel_path}")
        wb = openpyxl.load_workbook(self.excel_path)
        ws = wb['Factual Exhibits']
        
        logger.info(f"Worksheet: {ws.title}")
        logger.info(f"Current rows: {ws.max_row}")
        
        # Find insert position (after C-156, which is row 164)
        insert_row = 165
        
        # Verify C-156 is at row 164
        ref_cell = ws.cell(row=164, column=1).value
        if ref_cell != 'C-156':
            logger.warning(f"⚠️  Expected C-156 at row 164, found: {ref_cell}")
            logger.warning("   Searching for C-156...")
            
            # Search for C-156
            for row_idx in range(1, ws.max_row + 1):
                if ws.cell(row=row_idx, column=1).value == 'C-156':
                    insert_row = row_idx + 1
                    logger.info(f"   Found C-156 at row {row_idx}, inserting at row {insert_row}")
                    break
        
        logger.info(f"Inserting PHL documents starting at row: {insert_row}")
        
        # Column mapping (based on row 2 headers)
        col_map = {
            'reference': 1,           # A
            'date': 2,                # B
            'description': 3,         # C
            'trial_bundle_ref': 4,    # D
            'trial_bundle_index': 5,  # E
            'when_produced': 6,       # F
            'notes': 7                # G (NEW COLUMN)
        }
        
        # Add "Notes" header if column G is empty
        if ws.cell(row=2, column=7).value is None:
            ws.cell(row=2, column=7).value = "Notes"
            # Copy formatting from column F header
            source_cell = ws.cell(row=2, column=6)
            target_cell = ws.cell(row=2, column=7)
            if source_cell.font:
                target_cell.font = source_cell.font.copy()
            if source_cell.alignment:
                target_cell.alignment = source_cell.alignment.copy()
            if source_cell.border:
                target_cell.border = source_cell.border.copy()
            if source_cell.fill:
                target_cell.fill = source_cell.fill.copy()
        
        # Insert rows
        logger.info(f"Inserting {len(phl_docs)} rows...")
        ws.insert_rows(insert_row, len(phl_docs))
        
        # Get reference row for formatting (row 164 = C-156)
        reference_row = 164
        
        # Populate new rows
        for idx, doc in enumerate(tqdm(phl_docs, desc="Writing to Excel"), start=insert_row):
            # Write data
            ws.cell(row=idx, column=col_map['reference']).value = doc['reference']
            ws.cell(row=idx, column=col_map['date']).value = doc['date']
            ws.cell(row=idx, column=col_map['description']).value = doc['description']
            ws.cell(row=idx, column=col_map['trial_bundle_ref']).value = doc['trial_bundle_ref']
            ws.cell(row=idx, column=col_map['trial_bundle_index']).value = doc['trial_bundle_index_desc']
            ws.cell(row=idx, column=col_map['when_produced']).value = doc['when_produced']
            ws.cell(row=idx, column=col_map['notes']).value = doc['notes']
            
            # Copy formatting from reference row
            for col_idx in range(1, 8):
                source_cell = ws.cell(row=reference_row, column=col_idx)
                target_cell = ws.cell(row=idx, column=col_idx)
                
                # Copy styles
                if source_cell.font:
                    target_cell.font = source_cell.font.copy()
                if source_cell.alignment:
                    target_cell.alignment = source_cell.alignment.copy()
                if source_cell.border:
                    target_cell.border = source_cell.border.copy()
                if source_cell.fill:
                    target_cell.fill = source_cell.fill.copy()
                if source_cell.number_format:
                    target_cell.number_format = source_cell.number_format
        
        # Auto-adjust column G width for Notes
        ws.column_dimensions['G'].width = 30
        
        # Save output
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_filename = f"Claimants_Exhibits_with_PHL_{timestamp}.xlsx"
        output_path = self.output_folder / output_filename
        
        logger.info(f"Saving output: {output_path}")
        wb.save(output_path)
        wb.close()
        
        logger.info(f"✅ Excel file saved successfully")
        logger.info(f"   New total rows: {ws.max_row}")
        
        return output_path
    
    def generate_summary_report(self, output_path: Path):
        """
        Generate summary report of ingestion process.
        
        Args:
            output_path: Path to output Excel file
        """
        logger.info("\n" + "=" * 80)
        logger.info("SUMMARY REPORT")
        logger.info("=" * 80)
        
        total_docs = len(self.phl_documents)
        
        logger.info(f"\n📊 Documents Processed:")
        logger.info(f"   Total PHL documents added: {total_docs}")
        
        if total_docs > 0:
            first_phl = min(doc['phl_number'] for doc in self.phl_documents)
            last_phl = max(doc['phl_number'] for doc in self.phl_documents)
            logger.info(f"   Range: PHL_{first_phl:06d} to PHL_{last_phl:06d}")
        
        logger.info(f"\n⚠️  Issues Detected:")
        logger.info(f"   Duplicate documents: {len(self.duplicates)}")
        if self.duplicates:
            for dup in self.duplicates[:5]:  # Show first 5
                logger.info(f"      PHL_{dup['phl_number']:06d} (found in {len(dup['paths'])} locations)")
            if len(self.duplicates) > 5:
                logger.info(f"      ... and {len(self.duplicates) - 5} more")
        
        logger.info(f"   Metadata extraction failures: {len(self.missing_metadata)}")
        if self.missing_metadata:
            sample = self.missing_metadata[:5]
            logger.info(f"      PHL numbers: {', '.join(f'PHL_{n:06d}' for n in sample)}")
            if len(self.missing_metadata) > 5:
                logger.info(f"      ... and {len(self.missing_metadata) - 5} more")
        
        logger.info(f"\n📁 Output:")
        logger.info(f"   File: {output_path.name}")
        logger.info(f"   Location: {output_path.parent}")
        logger.info(f"   Size: {output_path.stat().st_size / 1024:.1f} KB")
        
        logger.info("\n✅ PHASE 1 COMPLETE")
        logger.info("=" * 80)
    
    def run(self) -> Path:
        """
        Execute complete PHL ingestion workflow.
        
        Returns:
            Path to output Excel file
        """
        try:
            # Phase 1: Scan directories
            self.phl_documents = self.scan_phl_directories()
            
            if not self.phl_documents:
                logger.error("ERROR: No PHL documents found!")
                return None
            
            # Phase 2: Update Excel
            output_path = self.append_to_excel(self.phl_documents)
            
            # Phase 3: Generate report
            self.generate_summary_report(output_path)
            
            return output_path
            
        except Exception as e:
            logger.error(f"❌ FATAL ERROR: {e}", exc_info=True)
            raise


def main():
    """Main execution function."""
    
    # Configuration
    EXCEL_PATH = r"C:\Users\JemAndrew\Downloads\Claimant's Exhibits Compared to Trial Bundle - Draft 2 - 12.10.25.xlsx"
    OUTPUT_FOLDER = r"C:\Users\JemAndrew\OneDrive - Velitor\Claude\litigation_analysis\cases\lismore_v_ph\analysis"
    BASE_PHL_PATH = r"C:\Users\JemAndrew\Velitor\Communication site - Documents\LIS1.1\73. Review of PHL Disclosures"
    
    # Create and run ingester
    ingester = PHLDocumentIngester(
        excel_path=EXCEL_PATH,
        output_folder=OUTPUT_FOLDER,
        base_phl_path=BASE_PHL_PATH
    )
    
    output_path = ingester.run()
    
    if output_path:
        print(f"\n🎉 SUCCESS! Output file: {output_path}")
        print(f"\nNext steps:")
        print(f"1. Review the output Excel file")
        print(f"2. Check 'Notes' column for any duplicates or issues")
        print(f"3. Proceed to Script 2 (document matching)")
    else:
        print(f"\n❌ FAILED - Check logs for details")


if __name__ == '__main__':
    main()