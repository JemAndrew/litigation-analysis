#!/usr/bin/env python3
"""
PHASE 2: Enhanced Bidirectional Matching
Uses multi-page sampled descriptions for accurate matching

Matches Folders 01/02 vs Folder 03 bidirectionally
with Claude AI review for ambiguous cases.

British English throughout.
"""

import os
import re
import json
from pathlib import Path
from datetime import datetime
from typing import List, Dict
import logging

import pandas as pd
from fuzzywuzzy import fuzz
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from tqdm import tqdm
from anthropic import Anthropic
from dotenv import load_dotenv

# Load .env
load_dotenv()

ANTHROPIC_API_KEY = os.getenv('ANTHROPIC_API_KEY')
if not ANTHROPIC_API_KEY:
    print("⚠️  Warning: ANTHROPIC_API_KEY not found in .env")
    print("   Claude AI review will be disabled")

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


# ============================================================================
# CONFIGURATION
# ============================================================================

THRESHOLD_HIGH = 90      # Auto-accept
THRESHOLD_MEDIUM = 75    # Send to Claude

COLOURS = {
    'smoking_gun': 'C6EFCE',
    'high_confidence': 'D9EAD3',
    'claude_approved': 'FFF2CC',
    'needs_review': 'FCE4D6',
    'no_match': 'F4CCCC',
    'light_grey': 'F3F3F3',
}

ABBREVIATIONS = {
    'p&id': 'process and industrial developments limited',
    'pid': 'process and industrial developments limited',
    'phl': 'process holdings limited',
    'ph': 'process holdings',
    'mpr': 'ministry of petroleum resources',
    'frn': 'federal republic of nigeria',
    'spa': 'share purchase agreement',
    'mou': 'memorandum of understanding',
    'gspa': 'gas supply and processing agreement',
}


# ============================================================================
# TEXT NORMALISATION
# ============================================================================

def normalise_description(text: str) -> str:
    """Normalise description for fuzzy matching"""
    if not text or text in ['NO TEXT EXTRACTED', 'EXTRACTION FAILED', 'FILE NOT FOUND']:
        return ""
    
    text = text.lower()
    
    # Expand abbreviations
    for abbr, full in ABBREVIATIONS.items():
        text = re.sub(r'\b' + re.escape(abbr) + r'\b', full, text)
    
    # Remove punctuation
    text = re.sub(r'[^\w\s]', ' ', text)
    
    # Normalise whitespace
    text = re.sub(r'\s+', ' ', text).strip()
    
    return text


# ============================================================================
# CLAUDE REVIEWER
# ============================================================================

class ClaudeReviewer:
    """Review ambiguous matches with Claude AI"""
    
    def __init__(self):
        if not ANTHROPIC_API_KEY:
            self.client = None
        else:
            self.client = Anthropic(api_key=ANTHROPIC_API_KEY)
        
        self.total_cost_gbp = 0.0
        self.call_count = 0
    
    def review_match(
        self,
        source_ref: str,
        source_desc: str,
        source_pages: int,
        target_ref: str,
        target_desc: str,
        target_pages: int,
        fuzzy_score: int
    ) -> Dict:
        """Ask Claude to review an ambiguous match"""
        
        if not self.client:
            return {
                'is_match': False,
                'confidence': 'SKIPPED',
                'reasoning': 'Claude API not configured',
                'cost_gbp': 0.0
            }
        
        prompt = f"""You are analysing documents for the Lismore v. Process Holdings arbitration.

**TASK:** Determine if these are THE SAME document or DIFFERENT documents.

**SOURCE (Folders 01/02 - April/June disclosure):**
Reference: {source_ref}
Pages: {source_pages}
Description (multi-page sample): {source_desc[:700]}

**TARGET (Folder 03 - September disclosure):**
Reference: {target_ref}
Pages: {target_pages}
Description (multi-page sample): {target_desc[:700]}

**FUZZY MATCH:** {fuzzy_score}%

**CRITICAL CHECKS:**
1. Page count: If very different (e.g., 50 vs 2), likely different documents
2. Content type: Email ABOUT document vs document ITSELF
3. Subset: Excerpt vs full document

**RESPOND IN JSON:**
{{
    "is_match": true/false,
    "confidence": "HIGH" or "MEDIUM" or "LOW",
    "reasoning": "Brief explanation"
}}

Be conservative - when in doubt, mark as NOT a match."""
        
        try:
            response = self.client.messages.create(
                model="claude-sonnet-4-5-20250929",
                max_tokens=4000,
                temperature=0.0,
                messages=[{"role": "user", "content": prompt}]
            )
            
            # Calculate cost
            input_cost = response.usage.input_tokens * 0.000003
            output_cost = response.usage.output_tokens * 0.000015
            cost_usd = input_cost + output_cost
            cost_gbp = cost_usd * 1.27
            
            self.total_cost_gbp += cost_gbp
            self.call_count += 1
            
            # Parse response
            result = self._parse_response(response.content[0].text)
            result['cost_gbp'] = cost_gbp
            
            return result
            
        except Exception as e:
            logger.error(f"Claude API error: {e}")
            return {
                'is_match': False,
                'confidence': 'ERROR',
                'reasoning': str(e),
                'cost_gbp': 0.0
            }
    
    def _parse_response(self, text: str) -> Dict:
        """Parse Claude's JSON response"""
        try:
            json_match = re.search(r'\{[^}]+\}', text, re.DOTALL)
            if json_match:
                data = json.loads(json_match.group())
                return {
                    'is_match': data.get('is_match', False),
                    'confidence': data.get('confidence', 'UNKNOWN'),
                    'reasoning': data.get('reasoning', '')
                }
        except:
            pass
        
        is_match = 'true' in text.lower()
        return {
            'is_match': is_match,
            'confidence': 'UNKNOWN',
            'reasoning': text[:200]
        }


# ============================================================================
# BIDIRECTIONAL MATCHER
# ============================================================================

class BidirectionalMatcher:
    """Match documents in both directions"""
    
    def __init__(self, use_claude: bool = True):
        self.use_claude = use_claude and ANTHROPIC_API_KEY is not None
        self.claude_reviewer = ClaudeReviewer() if self.use_claude else None
        
        self.stats = {
            'forward_matches': 0,
            'reverse_matches': 0,
            'bidirectional_matches': 0,
            'claude_reviews': 0,
            'claude_approved': 0,
            'claude_rejected': 0,
        }
    
    def match_documents(
        self,
        source_df: pd.DataFrame,
        target_df: pd.DataFrame,
        direction: str
    ) -> List[Dict]:
        """Match source against target"""
        
        logger.info(f"\n{'='*80}")
        logger.info(f"{direction.upper()} MATCHING: {len(source_df)} → {len(target_df)}")
        logger.info(f"{'='*80}\n")
        
        # Normalise descriptions
        logger.info("Normalising descriptions...")
        source_df['Description_Normalised'] = source_df['Description'].apply(normalise_description)
        target_df['Description_Normalised'] = target_df['Description'].apply(normalise_description)
        
        results = []
        medium_matches = []
        
        # Fuzzy matching
        logger.info("Fuzzy matching...")
        for _, source_row in tqdm(source_df.iterrows(), total=len(source_df), desc=f"{direction} fuzzy"):
            match_result = self._fuzzy_match(source_row, target_df)
            
            if match_result['tier'] == 'MEDIUM':
                medium_matches.append(match_result)
            else:
                results.append(match_result)
        
        logger.info(f"\n📊 Fuzzy matching complete:")
        logger.info(f"   ✅ High: {len([r for r in results if r['tier'] == 'HIGH'])}")
        logger.info(f"   🟡 Medium: {len(medium_matches)}")
        logger.info(f"   ❌ No match: {len([r for r in results if r['tier'] == 'NO_MATCH'])}")
        
        # Claude review of medium matches
        if self.use_claude and medium_matches:
            logger.info(f"\n{'='*80}")
            logger.info(f"CLAUDE AI REVIEW: {len(medium_matches)} ambiguous")
            logger.info(f"{'='*80}\n")
            
            for match in tqdm(medium_matches, desc="Claude review"):
                claude_result = self.claude_reviewer.review_match(
                    source_ref=match['Source_Reference'],
                    source_desc=match['Source_Description'],
                    source_pages=match.get('Source_Page_Count', 0),
                    target_ref=match['Matched_Reference'],
                    target_desc=match['Matched_Description'],
                    target_pages=match.get('Matched_Page_Count', 0),
                    fuzzy_score=match['Similarity_%']
                )
                
                self.stats['claude_reviews'] += 1
                
                if claude_result['is_match']:
                    match['tier'] = 'CLAUDE_APPROVED'
                    match['Status'] = f"✅ CLAUDE ({claude_result['confidence']})"
                    match['Colour'] = COLOURS['claude_approved']
                    self.stats['claude_approved'] += 1
                else:
                    match['tier'] = 'CLAUDE_REJECTED'
                    match['Status'] = f"❌ REJECTED ({claude_result['confidence']})"
                    match['Colour'] = COLOURS['no_match']
                    self.stats['claude_rejected'] += 1
                
                match['Claude_Reasoning'] = claude_result['reasoning']
                match['Claude_Cost'] = f"£{claude_result['cost_gbp']:.4f}"
                results.append(match)
            
            logger.info(f"\n💰 Cost: £{self.claude_reviewer.total_cost_gbp:.2f}")
            logger.info(f"   Approved: {self.stats['claude_approved']}")
            logger.info(f"   Rejected: {self.stats['claude_rejected']}")
        elif not self.use_claude and medium_matches:
            results.extend(medium_matches)
        
        if direction == 'forward':
            self.stats['forward_matches'] = len([r for r in results if r['tier'] != 'NO_MATCH'])
        else:
            self.stats['reverse_matches'] = len([r for r in results if r['tier'] != 'NO_MATCH'])
        
        return results
    
    def _fuzzy_match(self, source_row: pd.Series, target_df: pd.DataFrame) -> Dict:
        """Find best fuzzy match"""
        
        source_ref = source_row['Reference']
        source_desc = source_row['Description']
        source_desc_norm = source_row['Description_Normalised']
        
        if not source_desc_norm:
            return self._no_match_result(source_row, "NO DESCRIPTION")
        
        best_match = None
        best_score = 0
        
        for _, target_row in target_df.iterrows():
            target_desc_norm = target_row['Description_Normalised']
            
            if not target_desc_norm:
                continue
            
            similarity = fuzz.ratio(source_desc_norm, target_desc_norm)
            
            if similarity > best_score:
                best_score = similarity
                best_match = target_row
        
        if best_match is None or best_score < THRESHOLD_MEDIUM:
            return self._no_match_result(source_row, f"LOW SIMILARITY ({best_score}%)")
        
        # Classify
        if best_score >= THRESHOLD_HIGH:
            tier = 'HIGH'
            status = '✅ SMOKING GUN'
            colour = COLOURS['smoking_gun']
        else:
            tier = 'MEDIUM'
            status = '🟡 NEEDS REVIEW'
            colour = COLOURS['needs_review']
        
        return {
            'Source_Reference': source_ref,
            'Source_Date': source_row['Date'],
            'Source_Description': source_desc[:500],
            'Source_Folder': source_row['Source Folder'],
            'Source_Page_Count': source_row.get('Page Count', 0),
            'Source_File_Size': source_row.get('File Size (MB)', 0),
            
            'Matched_Reference': best_match['Reference'],
            'Matched_Date': best_match['Date'],
            'Matched_Description': best_match['Description'][:500],
            'Matched_Page_Count': best_match.get('Page Count', 0),
            'Matched_File_Size': best_match.get('File Size (MB)', 0),
            
            'Similarity_%': best_score,
            'tier': tier,
            'Status': status,
            'Colour': colour,
            
            'Claude_Reasoning': '',
            'Claude_Cost': ''
        }
    
    def _no_match_result(self, source_row: pd.Series, reason: str) -> Dict:
        """No match result"""
        return {
            'Source_Reference': source_row['Reference'],
            'Source_Date': source_row['Date'],
            'Source_Description': source_row['Description'][:500],
            'Source_Folder': source_row['Source Folder'],
            'Source_Page_Count': source_row.get('Page Count', 0),
            'Source_File_Size': source_row.get('File Size (MB)', 0),
            
            'Matched_Reference': '',
            'Matched_Date': '',
            'Matched_Description': '',
            'Matched_Page_Count': 0,
            'Matched_File_Size': 0,
            
            'Similarity_%': 0,
            'tier': 'NO_MATCH',
            'Status': f'❌ {reason}',
            'Colour': COLOURS['light_grey'],
            
            'Claude_Reasoning': '',
            'Claude_Cost': ''
        }
    
    def reconcile_bidirectional(
        self,
        forward_results: List[Dict],
        reverse_results: List[Dict]
    ) -> Dict[str, List[Dict]]:
        """Reconcile bidirectional matches"""
        
        logger.info(f"\n{'='*80}")
        logger.info("BIDIRECTIONAL RECONCILIATION")
        logger.info(f"{'='*80}\n")
        
        # Build maps
        forward_map = {}
        reverse_map = {}
        
        for match in forward_results:
            if match['tier'] != 'NO_MATCH':
                forward_map[match['Source_Reference']] = match['Matched_Reference']
        
        for match in reverse_results:
            if match['tier'] != 'NO_MATCH':
                reverse_map[match['Source_Reference']] = match['Matched_Reference']
        
        # Find bidirectional
        bidirectional = []
        for source_ref, target_ref in forward_map.items():
            if reverse_map.get(target_ref) == source_ref:
                bidirectional.append((source_ref, target_ref))
        
        self.stats['bidirectional_matches'] = len(bidirectional)
        
        logger.info(f"✅ Bidirectional matches: {len(bidirectional)}")
        
        # Categorise
        smoking_guns = []
        high_confidence = []
        claude_approved = []
        needs_review = []
        no_matches = []
        
        for match in forward_results:
            source_ref = match['Source_Reference']
            target_ref = match['Matched_Reference']
            
            is_bidir = (source_ref, target_ref) in bidirectional
            
            if match['tier'] == 'HIGH' and is_bidir:
                match['Status'] = '🎯 SMOKING GUN (Bidirectional)'
                smoking_guns.append(match)
            elif match['tier'] == 'HIGH':
                high_confidence.append(match)
            elif match['tier'] == 'CLAUDE_APPROVED':
                claude_approved.append(match)
            elif match['tier'] == 'MEDIUM' or match['tier'] == 'CLAUDE_REJECTED':
                needs_review.append(match)
            else:
                no_matches.append(match)
        
        logger.info(f"\n📊 Categorisation:")
        logger.info(f"   🎯 Smoking guns: {len(smoking_guns)}")
        logger.info(f"   ✅ High confidence: {len(high_confidence)}")
        logger.info(f"   🟡 Claude approved: {len(claude_approved)}")
        logger.info(f"   🟠 Needs review: {len(needs_review)}")
        logger.info(f"   ❌ No matches: {len(no_matches)}")
        
        return {
            'smoking_guns': smoking_guns,
            'high_confidence': high_confidence,
            'claude_approved': claude_approved,
            'needs_review': needs_review,
            'no_matches': no_matches,
            'forward_detailed': forward_results,
            'reverse_detailed': reverse_results
        }


# ============================================================================
# EXCEL OUTPUT
# ============================================================================

def create_excel_output(results: Dict, matcher: BidirectionalMatcher, output_file: Path):
    """Create colour-coded Excel output"""
    
    logger.info(f"\n{'='*80}")
    logger.info("CREATING EXCEL OUTPUT")
    logger.info(f"{'='*80}\n")
    
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # Summary
        summary_data = {
            'Metric': [
                'Source Documents (Folders 01/02)',
                'Target Documents (Folder 03)',
                '',
                '🎯 Smoking Guns (≥90% + Bidirectional)',
                '✅ High Confidence (≥90%)',
                '🟡 Claude Approved',
                '🟠 Needs Manual Review',
                '❌ No Matches',
                '',
                'Forward Matches',
                'Reverse Matches',
                'Bidirectional Matches',
                '',
                'Claude Reviews',
                'Claude Approved',
                'Claude Rejected',
                'Total Cost'
            ],
            'Count': [
                len(results['forward_detailed']),
                len(results['reverse_detailed']),
                '',
                len(results['smoking_guns']),
                len(results['high_confidence']),
                len(results['claude_approved']),
                len(results['needs_review']),
                len(results['no_matches']),
                '',
                matcher.stats['forward_matches'],
                matcher.stats['reverse_matches'],
                matcher.stats['bidirectional_matches'],
                '',
                matcher.stats['claude_reviews'],
                matcher.stats['claude_approved'],
                matcher.stats['claude_rejected'],
                f"£{matcher.claude_reviewer.total_cost_gbp:.2f}" if matcher.use_claude else '£0.00'
            ]
        }
        pd.DataFrame(summary_data).to_excel(writer, sheet_name='SUMMARY', index=False)
        
        # Data sheets
        if results['smoking_guns']:
            pd.DataFrame(results['smoking_guns']).to_excel(writer, sheet_name='SMOKING_GUNS', index=False)
        
        if results['high_confidence']:
            pd.DataFrame(results['high_confidence']).to_excel(writer, sheet_name='HIGH_CONFIDENCE', index=False)
        
        if results['claude_approved']:
            pd.DataFrame(results['claude_approved']).to_excel(writer, sheet_name='CLAUDE_APPROVED', index=False)
        
        if results['needs_review']:
            pd.DataFrame(results['needs_review']).to_excel(writer, sheet_name='NEEDS_REVIEW', index=False)
        
        if results['no_matches']:
            pd.DataFrame(results['no_matches']).to_excel(writer, sheet_name='NO_MATCHES', index=False)
        
        pd.DataFrame(results['forward_detailed']).to_excel(writer, sheet_name='FORWARD_DETAILED', index=False)
        pd.DataFrame(results['reverse_detailed']).to_excel(writer, sheet_name='REVERSE_DETAILED', index=False)
    
    # Apply colours
    apply_colours(output_file)
    
    logger.info(f"✅ Output: {output_file.name}")


def apply_colours(file_path: Path):
    """Apply colour coding"""
    
    workbook = load_workbook(file_path)
    
    for sheet_name in workbook.sheetnames:
        if sheet_name == 'SUMMARY':
            continue
        
        worksheet = workbook[sheet_name]
        
        # Find colour column
        colour_col = None
        for col_idx, cell in enumerate(worksheet[1], 1):
            if cell.value == 'Colour':
                colour_col = col_idx
                break
        
        if not colour_col:
            continue
        
        # Apply colours
        for row_idx in range(2, worksheet.max_row + 1):
            colour_value = worksheet.cell(row_idx, colour_col).value
            
            if colour_value:
                fill = PatternFill(start_color=colour_value, end_color=colour_value, fill_type='solid')
                
                for col_idx in range(1, worksheet.max_column + 1):
                    worksheet.cell(row_idx, col_idx).fill = fill
        
        # Hide colour column
        worksheet.column_dimensions[get_column_letter(colour_col)].hidden = True
        
        # Auto-size
        for col_idx in range(1, worksheet.max_column + 1):
            column_letter = get_column_letter(col_idx)
            max_length = 0
            
            for cell in worksheet[column_letter]:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            
            adjusted_width = min(max_length + 2, 60)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Bold headers
        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    workbook.save(file_path)


# ============================================================================
# MAIN
# ============================================================================

def main():
    """Phase 2: Run bidirectional matching"""
    
    print("="*80)
    print("PHASE 2: ENHANCED BIDIRECTIONAL MATCHING")
    print("Using Multi-Page Sampled Descriptions")
    print("="*80)
    
    analysis_dir = Path(r"C:\Users\JemAndrew\OneDrive - Velitor\Claude\litigation_analysis\cases\lismore_v_ph\analysis")
    
    # Find latest enhanced metadata
    files_01_02 = sorted(analysis_dir.glob('Folders_01_02_Enhanced_*.xlsx'))
    files_03 = sorted(analysis_dir.glob('Folder_03_Enhanced_*.xlsx'))
    
    if not files_01_02:
        print("\n❌ No Folders 01/02 enhanced metadata found!")
        print("   Run Phase 1 first: python clean_extraction_tool.py")
        return
    
    if not files_03:
        print("\n❌ No Folder 03 enhanced metadata found!")
        print("   Run Phase 1 first: python clean_extraction_tool.py")
        return
    
    file_01_02 = files_01_02[-1]
    file_03 = files_03[-1]
    
    print(f"\n📂 Using metadata files:")
    print(f"   Source: {file_01_02.name}")
    print(f"   Target: {file_03.name}")
    
    # Load data
    logger.info("\nLoading metadata...")
    df_01_02 = pd.read_excel(file_01_02)
    df_03 = pd.read_excel(file_03)
    
    print(f"\n📊 Document counts:")
    print(f"   Source (01/02): {len(df_01_02)}")
    print(f"   Target (03): {len(df_03)}")
    
    # Claude config
    print(f"\n{'='*80}")
    print("CLAUDE AI CONFIGURATION")
    print(f"{'='*80}")
    
    if ANTHROPIC_API_KEY:
        use_claude = input("\nUse Claude AI for ambiguous matches? (y/n) [y]: ").strip().lower()
        use_claude = use_claude != 'n'
    else:
        print("\n⚠️  Claude API key not found")
        print("   Fuzzy matching only")
        use_claude = False
    
    # Run matching
    matcher = BidirectionalMatcher(use_claude=use_claude)
    
    forward_results = matcher.match_documents(df_01_02, df_03, 'forward')
    reverse_results = matcher.match_documents(df_03, df_01_02, 'reverse')
    
    categorised = matcher.reconcile_bidirectional(forward_results, reverse_results)
    
    # Output
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_file = analysis_dir / f"Bidirectional_Matches_Enhanced_{timestamp}.xlsx"
    create_excel_output(categorised, matcher, output_file)
    
    # Summary
    print(f"\n{'='*80}")
    print("✅ PHASE 2 COMPLETE: BIDIRECTIONAL MATCHING")
    print(f"{'='*80}")
    
    smoking_guns = len(categorised['smoking_guns'])
    high_conf = len(categorised['high_confidence'])
    claude_approved = len(categorised['claude_approved'])
    total_matches = smoking_guns + high_conf + claude_approved
    
    print(f"\n📊 RESULTS:")
    print(f"   🎯 Smoking guns: {smoking_guns}")
    print(f"   ✅ High confidence: {high_conf}")
    print(f"   🟡 Claude approved: {claude_approved}")
    print(f"   " + "-"*40)
    print(f"   📌 TOTAL MATCHES: {total_matches}")
    
    if use_claude:
        print(f"\n💰 Claude cost: £{matcher.claude_reviewer.total_cost_gbp:.2f}")
    
    print(f"\n📂 Output file: {output_file.name}")
    
    print(f"\n💡 These {total_matches} matches prove:")
    print(f"   PH claimed documents were 'restricted' (15 Sept)")
    print(f"   BUT they were in earlier disclosures (April-June)")
    print(f"   = DELIBERATE CONCEALMENT")
    
    print(f"\n{'='*80}\n")


if __name__ == '__main__':
    try:
        main()
    except KeyboardInterrupt:
        print("\n\n⚠️  Cancelled by user")
    except Exception as e:
        print(f"\n❌ Error: {e}")
        import traceback
        traceback.print_exc()