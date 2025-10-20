#!/usr/bin/env python3
"""
Vision-First Autonomous Document Match Verifier

Let Claude see and think naturally. No rigid instructions.
Simple color-coded output showing confidence levels.

FIXED VERSION:
- Handles both PDFs AND Excel files
- Smart file finding (C-1 vs C-01 patterns)
- Detailed error logging
- Pre-flight document validation
- Memory management (explicit pixmap cleanup)
- PyMuPDF 1.18+ compatible (no quality parameter)

British English throughout.
"""

import os
import sys
from pathlib import Path
from typing import List, Dict, Optional, Tuple
import fitz  # PyMuPDF
import anthropic
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import logging
import time
import base64
from dotenv import load_dotenv
import json

# Load environment
load_dotenv()

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('autonomous_verification.log', encoding='utf-8'),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)


class AutonomousVisionVerifier:
    """Let Claude see documents and think autonomously"""
    
    def __init__(
        self,
        input_excel: str,
        c_exhibits_folder: str,
        phl_folder: str,
        restricted_folder: str,
        output_excel: str,
        max_cost_gbp: float = 100.0,
        pages_to_compare: int = 3,
        image_dpi: int = 150
    ):
        self.input_excel = Path(input_excel)
        self.c_exhibits_folder = Path(c_exhibits_folder)
        self.phl_folder = Path(phl_folder)
        self.restricted_folder = Path(restricted_folder)
        self.output_excel = Path(output_excel)
        self.max_cost_gbp = max_cost_gbp
        self.pages_to_compare = pages_to_compare
        self.image_dpi = image_dpi
        
        # API setup
        api_key = os.getenv('ANTHROPIC_API_KEY')
        if not api_key:
            raise ValueError("ANTHROPIC_API_KEY not found in .env")
        
        self.client = anthropic.Anthropic(api_key=api_key)
        
        # Cost tracking
        self.total_cost_gbp = 0.0
        self.api_call_count = 0
        self.INPUT_COST_PER_TOKEN = 0.000003
        self.OUTPUT_COST_PER_TOKEN = 0.000015
        self.USD_TO_GBP = 0.79
        
        # Results by confidence
        self.high_confidence: List[Dict] = []      # 90-100%
        self.medium_confidence: List[Dict] = []    # 70-89%
        self.low_confidence: List[Dict] = []       # 0-69%
        
        # All matches
        self.all_matches: List[Dict] = []
        
        # Checkpoint
        self.checkpoint_file = Path('vision_verification_checkpoint.json')
        self.processed_indices = set()
    
    def load_checkpoint(self) -> None:
        """Load checkpoint for resume capability"""
        if self.checkpoint_file.exists():
            try:
                with open(self.checkpoint_file, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    self.processed_indices = set(data.get('processed_indices', []))
                    self.total_cost_gbp = data.get('total_cost_gbp', 0.0)
                    self.high_confidence = data.get('high_confidence', [])
                    self.medium_confidence = data.get('medium_confidence', [])
                    self.low_confidence = data.get('low_confidence', [])
                    logger.info(f"Resuming: {len(self.processed_indices)} already processed")
            except Exception as e:
                logger.warning(f"Could not load checkpoint: {e}")
    
    def save_checkpoint(self) -> None:
        """Save checkpoint"""
        try:
            data = {
                'processed_indices': list(self.processed_indices),
                'total_cost_gbp': self.total_cost_gbp,
                'high_confidence': self.high_confidence,
                'medium_confidence': self.medium_confidence,
                'low_confidence': self.low_confidence,
                'timestamp': datetime.now().isoformat()
            }
            with open(self.checkpoint_file, 'w', encoding='utf-8') as f:
                json.dump(data, f, indent=2)
        except Exception as e:
            logger.warning(f"Could not save checkpoint: {e}")
    
    def load_input_matches(self) -> None:
        """Load all matches from input Excel"""
        logger.info(f"Loading matches from: {self.input_excel}")
        
        try:
            wb = openpyxl.load_workbook(self.input_excel, data_only=True)
            ws = wb['Cross-Match Analysis']
            
            for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row or not row[0]:
                    continue
                
                self.all_matches.append({
                    'index': idx,
                    'ref_a': str(row[0]).strip() if row[0] else '',
                    'date_a': str(row[1]).strip() if row[1] else '',
                    'desc_a': str(row[2]).strip() if row[2] else '',
                    'ref_b': str(row[3]).strip() if row[3] else '',
                    'date_b': str(row[4]).strip() if row[4] else '',
                    'desc_b': str(row[5]).strip() if row[5] else ''
                })
            
            wb.close()
            logger.info(f"Loaded {len(self.all_matches)} matches")
            
        except Exception as e:
            logger.error(f"Error loading input: {e}")
            raise
    
    def find_document(self, reference: str) -> Optional[Path]:
        """
        Find document by reference - handles PDFs AND Excel files
        
        Folder 01: C-XX files (PDFs and Excel)
        Folder 02: PHL_XXXXXX files only  
        Folder 03: Everything else (H1_, L1_, M_, F6_, A_, B_, etc.)
        
        Handles both C-1 and C-01 naming conventions
        Supports: .pdf, .xlsx, .xls
        """
        try:
            if reference.startswith('C-'):
                # Claimant's exhibits (Folder 01)
                # Try multiple patterns and file types
                
                # Get the number part
                number_part = reference.split('-')[1] if '-' in reference else None
                
                # Try different extensions
                extensions = ['.pdf', '.xlsx', '.xls']
                
                for ext in extensions:
                    # Pattern 1: Exact match (C-135 → C-135.pdf)
                    path = self.c_exhibits_folder / f"{reference}{ext}"
                    if path.exists():
                        return path
                    
                    # Pattern 2: Add leading zero for single/double digits
                    if number_part and number_part.isdigit():
                        number = int(number_part)
                        
                        # Try with leading zeros
                        if number < 10:
                            # Single digit: C-1 → C-01
                            zero_padded = f"C-{number:02d}{ext}"
                            path = self.c_exhibits_folder / zero_padded
                            if path.exists():
                                logger.debug(f"  Found {reference} as {zero_padded}")
                                return path
                        
                        elif number < 100:
                            # Double digit: might also need padding in some systems
                            zero_padded = f"C-{number:03d}{ext}"
                            path = self.c_exhibits_folder / zero_padded
                            if path.exists():
                                logger.debug(f"  Found {reference} as {zero_padded}")
                                return path
                    
                    # Pattern 3: Remove leading zeros (C-01 → C-1)
                    if number_part and number_part.isdigit():
                        no_zeros = f"C-{int(number_part)}{ext}"
                        path = self.c_exhibits_folder / no_zeros
                        if path.exists():
                            logger.debug(f"  Found {reference} as {no_zeros}")
                            return path
                
                # Pattern 4: Recursive search for all extensions
                for ext in extensions:
                    matches = list(self.c_exhibits_folder.rglob(f"{reference}{ext}"))
                    if matches:
                        return matches[0]
                    
                    # Try with leading zeros in recursive search
                    if number_part and number_part.isdigit():
                        number = int(number_part)
                        if number < 10:
                            zero_padded = f"C-{number:02d}{ext}"
                            matches = list(self.c_exhibits_folder.rglob(zero_padded))
                            if matches:
                                logger.debug(f"  Found {reference} as {zero_padded} (recursive)")
                                return matches[0]
                        elif number < 100:
                            zero_padded = f"C-{number:03d}{ext}"
                            matches = list(self.c_exhibits_folder.rglob(zero_padded))
                            if matches:
                                logger.debug(f"  Found {reference} as {zero_padded} (recursive)")
                                return matches[0]
                
                return None
            
            elif reference.startswith('PHL_'):
                # PHL disclosure (Folder 02) - usually PDFs only
                matches = list(self.phl_folder.rglob(f"{reference}.pdf"))
                return matches[0] if matches else None
            
            else:
                # Everything else goes to Folder 03 (restricted/trial bundle)
                # H1_, H7_, L1_, M_, F6_, A_, B_, etc.
                path = self.restricted_folder / f"{reference}.pdf"
                if path.exists():
                    return path
                # Recursive search if not in root
                matches = list(self.restricted_folder.rglob(f"{reference}.pdf"))
                return matches[0] if matches else None
                
        except Exception as e:
            logger.error(f"Error finding {reference}: {e}")
            return None
    
    def pdf_to_images(self, pdf_path: Path) -> Tuple[List[bytes], str]:
        """
        Convert first N pages of PDF to high-quality images.
        WITH DETAILED ERROR LOGGING!
        
        Returns: (list of JPEG bytes, notes)
        """
        images = []
        notes = []
        
        try:
            logger.debug(f"     Opening: {pdf_path.name}")
            doc = fitz.open(pdf_path)
            page_count = len(doc)
            pages_to_extract = min(page_count, self.pages_to_compare)
            
            logger.debug(f"     Pages: {page_count}, extracting: {pages_to_extract}")
            
            # Check if encrypted
            if doc.is_encrypted:
                error_msg = "PDF is encrypted/password-protected"
                logger.error(f"     ❌ {error_msg}")
                doc.close()
                return [], f"ERROR: {error_msg}"
            
            if page_count > self.pages_to_compare:
                notes.append(f"Showing first {self.pages_to_compare} of {page_count} pages")
            
            # Extract pages
            for page_num in range(pages_to_extract):
                try:
                    logger.debug(f"     Converting page {page_num + 1}/{pages_to_extract}...")
                    page = doc[page_num]
                    
                    # Render to pixmap (high quality)
                    pix = page.get_pixmap(dpi=self.image_dpi)
                    
                    # Convert to JPEG (NO quality parameter in PyMuPDF 1.18+)
                    img_bytes = pix.tobytes("jpeg")
                    
                    # Free memory immediately
                    del pix
                    
                    # Check size
                    if len(img_bytes) > 8_000_000:  # 8MB API limit
                        logger.warning(f"     ⚠️  Page {page_num + 1}: Large ({len(img_bytes)/1e6:.1f}MB), reducing DPI")
                        notes.append(f"Page {page_num + 1}: Large image, reduced DPI")
                        
                        # Retry with lower DPI
                        pix = page.get_pixmap(dpi=100)
                        img_bytes = pix.tobytes("jpeg")
                        del pix
                        
                        logger.debug(f"     ✅ Reduced to {len(img_bytes)/1e6:.1f}MB")
                    
                    images.append(img_bytes)
                    logger.debug(f"     ✅ Page {page_num + 1}: {len(img_bytes)/1e6:.2f}MB")
                    
                except Exception as e:
                    error_msg = f"Page {page_num + 1}: {type(e).__name__} - {str(e)}"
                    logger.error(f"     ❌ {error_msg}")
                    notes.append(error_msg)
                    continue
            
            doc.close()
            
            # Final status
            if images:
                logger.info(f"     ✅ Extracted {len(images)}/{pages_to_extract} pages successfully")
                notes_str = "; ".join(notes) if notes else f"Extracted {len(images)} pages"
            else:
                logger.error(f"     ❌ FAILED to extract ANY images from {pdf_path.name}")
                if notes:
                    logger.error(f"     Errors: {'; '.join(notes)}")
                notes_str = f"ERROR: No images extracted. {'; '.join(notes)}"
            
            return images, notes_str
            
        except Exception as e:
            error_msg = f"ERROR opening PDF: {type(e).__name__} - {str(e)}"
            logger.error(f"     ❌ {error_msg}")
            return [], error_msg
    
    def excel_to_images(self, excel_path: Path) -> Tuple[List[bytes], str]:
        """
        Convert Excel file to images for Claude Vision
        
        NOTE: Currently returns empty - Excel visual comparison not implemented
        Returns: (list of JPEG bytes, notes)
        """
        images = []
        notes = []
        
        try:
            logger.info(f"     Excel file detected: {excel_path.name}")
            
            # Load workbook to get basic info
            wb = openpyxl.load_workbook(excel_path, data_only=True)
            sheet_count = len(wb.sheetnames)
            
            logger.warning(f"     ⚠️  Excel files cannot be visually compared")
            logger.warning(f"     Sheets: {', '.join(wb.sheetnames)}")
            
            notes.append(f"Excel file with {sheet_count} sheet(s): {', '.join(wb.sheetnames)}")
            notes.append("Visual comparison not supported - manual review required")
            
            wb.close()
            
            return [], f"Excel file ({sheet_count} sheets) - cannot compare visually"
            
        except Exception as e:
            error_msg = f"ERROR reading Excel: {type(e).__name__} - {str(e)}"
            logger.error(f"     ❌ {error_msg}")
            return [], error_msg
    
    def validate_all_documents(self) -> Tuple[List[str], List[str]]:
        """
        Pre-flight check: Verify all documents exist before starting
        
        Returns: (found_refs, missing_refs)
        """
        logger.info("\n" + "="*70)
        logger.info("PRE-FLIGHT CHECK: Validating document references...")
        logger.info("="*70)
        
        found = []
        missing = []
        excel_files = []
        
        # Get all unique references
        all_refs = set()
        for match in self.all_matches:
            all_refs.add(match['ref_a'])
            all_refs.add(match['ref_b'])
        
        logger.info(f"Checking {len(all_refs)} unique document references...\n")
        
        for ref in sorted(all_refs):
            path = self.find_document(ref)
            if path:
                found.append(ref)
                if path.suffix.lower() in ['.xlsx', '.xls']:
                    excel_files.append(ref)
                    logger.debug(f"  📊 {ref}: {path} (EXCEL)")
                else:
                    logger.debug(f"  ✅ {ref}: {path}")
            else:
                missing.append(ref)
                logger.warning(f"  ❌ {ref}: NOT FOUND")
        
        logger.info(f"\n✅ Found: {len(found)}/{len(all_refs)} documents")
        
        if excel_files:
            logger.warning(f"📊 Excel files: {len(excel_files)} (cannot compare visually)")
            logger.warning(f"   Examples: {', '.join(excel_files[:5])}")
            if len(excel_files) > 5:
                logger.warning(f"   ... and {len(excel_files) - 5} more")
        
        if missing:
            logger.warning(f"⚠️  Missing: {len(missing)} documents")
            logger.warning(f"\nFirst 10 missing references:")
            for ref in missing[:10]:
                logger.warning(f"  - {ref}")
            if len(missing) > 10:
                logger.warning(f"  ... and {len(missing) - 10} more")
        
        logger.info("="*70 + "\n")
        
        return found, missing
    
    def verify_with_claude(
        self,
        ref_a: str,
        images_a: List[bytes],
        ref_b: str,
        images_b: List[bytes]
    ) -> Tuple[str, int, str, float]:
        """
        Let Claude see and compare documents autonomously.
        
        Returns: (verdict, confidence, reasoning, cost_gbp)
        """
        
        # Build content - images first, then simple prompt
        content = []
        
        # Simple, open-ended prompt
        prompt = f"""You are comparing two legal documents to determine if they are the same document.

DOCUMENT A: {ref_a}
DOCUMENT B: {ref_b}

Below are images of both documents.

Compare them carefully and tell me:

1. Are they the same document? (MATCH or NO MATCH)
2. How confident are you? (0-100%)
3. Why? (Brief explanation)

Respond in this format:
VERDICT: [MATCH or NO MATCH]
CONFIDENCE: [0-100]
REASON: [Brief explanation of your reasoning]

Trust your judgment and be thorough."""
        
        content.append({"type": "text", "text": prompt})
        
        # Add Document A images
        if images_a:
            content.append({"type": "text", "text": f"\n--- DOCUMENT A: {ref_a} ---"})
            for img_bytes in images_a:
                content.append({
                    "type": "image",
                    "source": {
                        "type": "base64",
                        "media_type": "image/jpeg",
                        "data": base64.b64encode(img_bytes).decode('utf-8')
                    }
                })
        else:
            content.append({"type": "text", "text": f"\n[Document A: Could not extract images]"})
        
        # Add Document B images
        if images_b:
            content.append({"type": "text", "text": f"\n--- DOCUMENT B: {ref_b} ---"})
            for img_bytes in images_b:
                content.append({
                    "type": "image",
                    "source": {
                        "type": "base64",
                        "media_type": "image/jpeg",
                        "data": base64.b64encode(img_bytes).decode('utf-8')
                    }
                })
        else:
            content.append({"type": "text", "text": f"\n[Document B: Could not extract images]"})
        
        try:
            response = self.client.messages.create(
                model="claude-sonnet-4-5-20250929",
                max_tokens=1500,
                temperature=0,
                messages=[{"role": "user", "content": content}]
            )
            
            response_text = response.content[0].text.strip()
            
            # Parse response
            verdict = "NO MATCH"
            confidence = 50
            reasoning = ""
            
            for line in response_text.split('\n'):
                line = line.strip()
                
                if line.startswith('VERDICT:'):
                    verdict_text = line.replace('VERDICT:', '').strip().upper()
                    verdict = "MATCH" if 'MATCH' in verdict_text and 'NO MATCH' not in verdict_text else "NO MATCH"
                
                elif line.startswith('CONFIDENCE:'):
                    try:
                        confidence = int(''.join(filter(str.isdigit, line)))
                        confidence = max(0, min(100, confidence))
                    except:
                        confidence = 50
                
                elif line.startswith('REASON:'):
                    reasoning = line.replace('REASON:', '').strip()
            
            # If reason spans multiple lines
            if 'REASON:' in response_text:
                reasoning_parts = response_text.split('REASON:', 1)
                if len(reasoning_parts) > 1:
                    reasoning = reasoning_parts[1].strip()
            
            if not reasoning:
                reasoning = response_text[:300]
            
            # Calculate cost
            input_tokens = response.usage.input_tokens
            output_tokens = response.usage.output_tokens
            cost_usd = (input_tokens * self.INPUT_COST_PER_TOKEN) + \
                      (output_tokens * self.OUTPUT_COST_PER_TOKEN)
            cost_gbp = cost_usd * self.USD_TO_GBP
            
            self.total_cost_gbp += cost_gbp
            self.api_call_count += 1
            
            logger.info(f"     {verdict} | Confidence: {confidence}% | Cost: £{cost_gbp:.4f}")
            
            return verdict, confidence, reasoning, cost_gbp
            
        except Exception as e:
            logger.error(f"Claude API error: {e}")
            return "ERROR", 0, f"API Error: {str(e)}", 0.0
    
    def process_all_matches(self) -> None:
        """Process all matches autonomously"""
        
        logger.info("="*70)
        logger.info("VISION-FIRST AUTONOMOUS VERIFICATION")
        logger.info("="*70)
        logger.info(f"Total matches: {len(self.all_matches)}")
        logger.info(f"Cost limit: £{self.max_cost_gbp}")
        logger.info(f"Pages per document: {self.pages_to_compare}")
        logger.info(f"Image DPI: {self.image_dpi}")
        logger.info("="*70)
        
        for idx, match in enumerate(self.all_matches, start=1):
            # Skip if processed
            if match['index'] in self.processed_indices:
                logger.info(f"[{idx}/{len(self.all_matches)}] SKIPPED: Already done")
                continue
            
            # Check cost limit
            if self.total_cost_gbp >= self.max_cost_gbp:
                logger.warning(f"COST LIMIT REACHED: £{self.total_cost_gbp:.2f}")
                logger.warning(f"Stopping at {idx-1}/{len(self.all_matches)}")
                break
            
            logger.info(f"\n[{idx}/{len(self.all_matches)}] COMPARING:")
            logger.info(f"  {match['ref_a']} <-> {match['ref_b']}")
            
            # Find documents
            path_a = self.find_document(match['ref_a'])
            path_b = self.find_document(match['ref_b'])
            
            if not path_a:
                logger.warning(f"  ❌ NOT FOUND: {match['ref_a']}")
                result = {
                    **match,
                    'verdict': 'ERROR',
                    'confidence': 0,
                    'reasoning': f"Document not found: {match['ref_a']}",
                    'cost': 0.0
                }
                self.low_confidence.append(result)
                self.processed_indices.add(match['index'])
                continue
            
            if not path_b:
                logger.warning(f"  ❌ NOT FOUND: {match['ref_b']}")
                result = {
                    **match,
                    'verdict': 'ERROR',
                    'confidence': 0,
                    'reasoning': f"Document not found: {match['ref_b']}",
                    'cost': 0.0
                }
                self.low_confidence.append(result)
                self.processed_indices.add(match['index'])
                continue
            
            # Convert to images (handle both PDF and Excel)
            logger.info(f"  Converting {path_a.name} to images...")
            
            if path_a.suffix.lower() == '.pdf':
                images_a, notes_a = self.pdf_to_images(path_a)
            elif path_a.suffix.lower() in ['.xlsx', '.xls']:
                images_a, notes_a = self.excel_to_images(path_a)
            else:
                images_a, notes_a = [], f"Unsupported file type: {path_a.suffix}"
            
            logger.info(f"  Converting {path_b.name} to images...")
            
            if path_b.suffix.lower() == '.pdf':
                images_b, notes_b = self.pdf_to_images(path_b)
            elif path_b.suffix.lower() in ['.xlsx', '.xls']:
                images_b, notes_b = self.excel_to_images(path_b)
            else:
                images_b, notes_b = [], f"Unsupported file type: {path_b.suffix}"
            
            # Check if extraction succeeded
            if not images_a or not images_b:
                # Check if it's because of Excel files
                is_excel_a = path_a.suffix.lower() in ['.xlsx', '.xls']
                is_excel_b = path_b.suffix.lower() in ['.xlsx', '.xls']
                
                if is_excel_a or is_excel_b:
                    logger.warning(f"  📊 EXCEL FILE DETECTED - SKIPPING")
                    if is_excel_a:
                        logger.warning(f"     {match['ref_a']}: {path_a.name}")
                    if is_excel_b:
                        logger.warning(f"     {match['ref_b']}: {path_b.name}")
                    
                    result = {
                        **match,
                        'verdict': 'CANNOT_VERIFY',
                        'confidence': 0,
                        'reasoning': f"Excel file(s) detected - manual review required. A: {notes_a}; B: {notes_b}",
                        'cost': 0.0
                    }
                    self.low_confidence.append(result)
                    self.processed_indices.add(match['index'])
                    continue
                
                # Otherwise, it's a genuine PDF extraction failure
                if not images_a:
                    logger.error(f"  ❌ FAILED: {match['ref_a']}")
                    logger.error(f"     Reason: {notes_a}")
                if not images_b:
                    logger.error(f"  ❌ FAILED: {match['ref_b']}")
                    logger.error(f"     Reason: {notes_b}")
                
                result = {
                    **match,
                    'verdict': 'ERROR',
                    'confidence': 0,
                    'reasoning': f"Image extraction failed. A: {notes_a}; B: {notes_b}",
                    'cost': 0.0
                }
                self.low_confidence.append(result)
                self.processed_indices.add(match['index'])
                continue
            
            # Let Claude see and compare
            logger.info(f"  Sending to Claude Vision (autonomous analysis)...")
            logger.info(f"  Doc A: {len(images_a)} images | Doc B: {len(images_b)} images")
            
            verdict, confidence, reasoning, cost = self.verify_with_claude(
                match['ref_a'], images_a,
                match['ref_b'], images_b
            )
            
            # Create result
            result = {
                **match,
                'verdict': verdict,
                'confidence': confidence,
                'reasoning': reasoning,
                'cost': cost
            }
            
            # Categorise by confidence
            if confidence >= 90:
                self.high_confidence.append(result)
                logger.info(f"  ✅ HIGH CONFIDENCE: {verdict} ({confidence}%)")
            elif confidence >= 70:
                self.medium_confidence.append(result)
                logger.info(f"  ✅ MEDIUM CONFIDENCE: {verdict} ({confidence}%)")
            else:
                self.low_confidence.append(result)
                logger.info(f"  ⚠️  LOW CONFIDENCE: {verdict} ({confidence}%)")
            
            # Mark processed
            self.processed_indices.add(match['index'])
            
            # Save checkpoint every 10
            if idx % 10 == 0:
                self.save_checkpoint()
                logger.info(f"  💾 Checkpoint saved | Total cost: £{self.total_cost_gbp:.2f}")
            
            # Rate limiting
            time.sleep(0.5)
        
        logger.info("\n" + "="*70)
        logger.info("VERIFICATION COMPLETE")
        logger.info("="*70)
        logger.info(f"Processed: {len(self.processed_indices)}/{len(self.all_matches)}")
        logger.info(f"High confidence (90%+): {len(self.high_confidence)}")
        logger.info(f"Medium confidence (70-89%): {len(self.medium_confidence)}")
        logger.info(f"Low confidence (<70%): {len(self.low_confidence)}")
        logger.info(f"Total cost: £{self.total_cost_gbp:.2f}")
    
    def get_confidence_color(self, confidence: int) -> PatternFill:
        """Get colour for confidence level"""
        if confidence >= 90:
            # Dark green for high confidence
            return PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
        elif confidence >= 80:
            # Light green
            return PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
        elif confidence >= 70:
            # Yellow
            return PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        elif confidence >= 50:
            # Orange
            return PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        else:
            # Red for low confidence
            return PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    
    def create_output_excel(self) -> Path:
        """Create clean Excel with 3 sheets, colour-coded confidence"""
        
        logger.info("Creating output Excel with colour-coded confidence...")
        
        try:
            wb = openpyxl.Workbook()
            
            # Remove default sheet
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
            
            # Define simple headers
            headers = [
                'Reference A',
                'Description A',
                'Reference B',
                'Description B',
                'Match Confidence %',
                'Brief Reasoning'
            ]
            
            # Header style
            header_font = Font(bold=True, size=11, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            # Sheet 1: High Confidence (90-100%)
            ws1 = wb.create_sheet("High Confidence (90-100%)")
            ws1.append(headers)
            for cell in ws1[1]:
                cell.font = header_font
                cell.fill = header_fill
            
            for match in self.high_confidence:
                row_num = ws1.max_row + 1
                ws1.append([
                    match['ref_a'],
                    match['desc_a'],
                    match['ref_b'],
                    match['desc_b'],
                    match['confidence'],
                    match['reasoning']
                ])
                
                # Colour code confidence cell
                conf_cell = ws1.cell(row=row_num, column=5)
                conf_cell.fill = self.get_confidence_color(match['confidence'])
                conf_cell.font = Font(bold=True)
                conf_cell.alignment = Alignment(horizontal='center')
            
            # Sheet 2: Medium Confidence (70-89%)
            ws2 = wb.create_sheet("Medium Confidence (70-89%)")
            ws2.append(headers)
            for cell in ws2[1]:
                cell.font = header_font
                cell.fill = header_fill
            
            for match in self.medium_confidence:
                row_num = ws2.max_row + 1
                ws2.append([
                    match['ref_a'],
                    match['desc_a'],
                    match['ref_b'],
                    match['desc_b'],
                    match['confidence'],
                    match['reasoning']
                ])
                
                conf_cell = ws2.cell(row=row_num, column=5)
                conf_cell.fill = self.get_confidence_color(match['confidence'])
                conf_cell.font = Font(bold=True)
                conf_cell.alignment = Alignment(horizontal='center')
            
            # Sheet 3: Low Confidence (<70%)
            ws3 = wb.create_sheet("Low Confidence (<70%)")
            ws3.append(headers)
            for cell in ws3[1]:
                cell.font = header_font
                cell.fill = header_fill
            
            for match in self.low_confidence:
                row_num = ws3.max_row + 1
                ws3.append([
                    match['ref_a'],
                    match['desc_a'],
                    match['ref_b'],
                    match['desc_b'],
                    match['confidence'],
                    match['reasoning']
                ])
                
                conf_cell = ws3.cell(row=row_num, column=5)
                conf_cell.fill = self.get_confidence_color(match['confidence'])
                conf_cell.font = Font(bold=True)
                conf_cell.alignment = Alignment(horizontal='center')
            
            # Set column widths for all sheets
            for ws in [ws1, ws2, ws3]:
                ws.column_dimensions['A'].width = 18  # Ref A
                ws.column_dimensions['B'].width = 50  # Desc A
                ws.column_dimensions['C'].width = 18  # Ref B
                ws.column_dimensions['D'].width = 50  # Desc B
                ws.column_dimensions['E'].width = 20  # Confidence (colour-coded)
                ws.column_dimensions['F'].width = 60  # Reasoning
                
                # Freeze top row
                ws.freeze_panes = 'A2'
            
            # Save
            wb.save(self.output_excel)
            wb.close()
            
            logger.info(f"✅ Output created: {self.output_excel}")
            logger.info(f"  Sheet 1 (High):   {len(self.high_confidence)} matches")
            logger.info(f"  Sheet 2 (Medium): {len(self.medium_confidence)} matches")
            logger.info(f"  Sheet 3 (Low):    {len(self.low_confidence)} matches")
            
            return self.output_excel
            
        except Exception as e:
            logger.error(f"Error creating Excel: {e}")
            raise
    
    def run(self) -> Path:
        """Execute full autonomous verification"""
        
        try:
            # Load checkpoint
            self.load_checkpoint()
            
            # Load matches
            self.load_input_matches()
            
            # Pre-flight check
            found, missing = self.validate_all_documents()
            
            if missing:
                print(f"\n⚠️  WARNING: {len(missing)} documents not found!")
                response = input("Continue anyway? (y/n): ")
                if response.lower() != 'y':
                    logger.info("❌ Cancelled by user")
                    return None
            
            # Process all
            self.process_all_matches()
            
            # Create output
            output_path = self.create_output_excel()
            
            # Clean up checkpoint
            if self.checkpoint_file.exists():
                self.checkpoint_file.unlink()
            
            return output_path
            
        except Exception as e:
            logger.error(f"Fatal error: {e}")
            try:
                self.save_checkpoint()
            except:
                pass
            raise


def main():
    """Main execution"""
    
    input_excel = r"C:\Users\JemAndrew\OneDrive - Velitor\Claude\litigation_analysis\cases\lismore_v_ph\analysis\Cross_Match_Analysis_Results.xlsx"
    
    c_exhibits_folder = r"C:\Users\JemAndrew\Velitor\Communication site - Documents\LIS1.1\73. Review of PHL Disclosures\01. Claimant's Factual Exhibits"
    
    phl_folder = r"C:\Users\JemAndrew\Velitor\Communication site - Documents\LIS1.1\73. Review of PHL Disclosures\02. PHL Disclosure prior to 23 June"
    
    restricted_folder = r"C:\Users\JemAndrew\Velitor\Communication site - Documents\LIS1.1\73. Review of PHL Disclosures\03. Erroneously restricted documents"
    
    output_excel = r"C:\Users\JemAndrew\OneDrive - Velitor\Claude\litigation_analysis\cases\lismore_v_ph\analysis\Autonomous_Match_Results.xlsx"
    
    # Create verifier
    verifier = AutonomousVisionVerifier(
        input_excel=input_excel,
        c_exhibits_folder=c_exhibits_folder,
        phl_folder=phl_folder,
        restricted_folder=restricted_folder,
        output_excel=output_excel,
        max_cost_gbp=100.0,     # Adjust as needed
        pages_to_compare=3,      # First 3 pages
        image_dpi=150            # High quality
    )
    
    # Run
    output_path = verifier.run()
    
    if output_path:
        print("\n" + "="*70)
        print("✅ AUTONOMOUS VERIFICATION COMPLETE!")
        print("="*70)
        print(f"Output: {output_path}")
        print(f"\nResults by Confidence:")
        print(f"  90-100% (High):   {len(verifier.high_confidence)} matches")
        print(f"  70-89% (Medium):  {len(verifier.medium_confidence)} matches")
        print(f"  0-69% (Low):      {len(verifier.low_confidence)} matches")
        print(f"\nTotal processed: {len(verifier.processed_indices)}/{len(verifier.all_matches)}")
        print(f"Total cost: £{verifier.total_cost_gbp:.2f}")
        print(f"API calls: {verifier.api_call_count}")
        print("="*70)


if __name__ == "__main__":
    main()