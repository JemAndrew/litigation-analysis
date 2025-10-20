#!/usr/bin/env python3
"""
Add VERDICT column to existing Excel output

Reads the checkpoint JSON (which has all verdicts) and updates the Excel
WITHOUT re-running the expensive API calls!

British English throughout.
"""

import json
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


def add_verdict_column():
    """Add verdict column to existing Excel"""
    
    print("="*70)
    print("ADDING VERDICT COLUMN TO EXCEL")
    print("="*70)
    print()
    
    # Paths
    checkpoint_file = Path('vision_verification_checkpoint.json')
    excel_file = Path(r"C:\Users\JemAndrew\OneDrive - Velitor\Claude\litigation_analysis\cases\lismore_v_ph\analysis\Autonomous_Match_Results.xlsx")
    output_file = Path(r"C:\Users\JemAndrew\OneDrive - Velitor\Claude\litigation_analysis\cases\lismore_v_ph\analysis\Autonomous_Match_Results_WITH_VERDICT.xlsx")
    
    # Check if checkpoint exists
    if not checkpoint_file.exists():
        print("❌ Checkpoint file not found!")
        print(f"   Looking for: {checkpoint_file}")
        print()
        print("Alternative: Extract from log file")
        return extract_from_log()
    
    # Load checkpoint data
    print(f"📂 Loading checkpoint: {checkpoint_file.name}")
    with open(checkpoint_file, 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    high_conf = data.get('high_confidence', [])
    medium_conf = data.get('medium_confidence', [])
    low_conf = data.get('low_confidence', [])
    
    print(f"   High confidence: {len(high_conf)}")
    print(f"   Medium confidence: {len(medium_conf)}")
    print(f"   Low confidence: {len(low_conf)}")
    print()
    
    # Load Excel
    print(f"📊 Loading Excel: {excel_file.name}")
    wb = openpyxl.load_workbook(excel_file)
    
    # Process each sheet
    for sheet_name, results in [
        ("High Confidence (90-100%)", high_conf),
        ("Medium Confidence (70-89%)", medium_conf),
        ("Low Confidence (<70%)", low_conf)
    ]:
        if sheet_name not in wb.sheetnames:
            print(f"   ⚠️  Sheet not found: {sheet_name}")
            continue
        
        ws = wb[sheet_name]
        
        print(f"\n   Processing: {sheet_name}")
        
        # Insert new column after "Reference B" (column D)
        # New headers will be: Ref A, Desc A, Ref B, Desc B, VERDICT, Confidence, Reasoning
        ws.insert_cols(5)  # Insert at column E (pushes Confidence and Reasoning right)
        
        # Update header
        ws.cell(row=1, column=5).value = "VERDICT"
        ws.cell(row=1, column=5).font = Font(bold=True, size=11, color="FFFFFF")
        ws.cell(row=1, column=5).fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        ws.cell(row=1, column=5).alignment = Alignment(horizontal='center')
        
        # Add verdict for each row
        for idx, result in enumerate(results, start=2):
            verdict = result.get('verdict', 'UNKNOWN')
            
            # Set verdict
            cell = ws.cell(row=idx, column=5)
            cell.value = verdict
            cell.alignment = Alignment(horizontal='center')
            cell.font = Font(bold=True, size=11)
            
            # Colour code verdict
            if verdict == "MATCH":
                cell.fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
                cell.font = Font(bold=True, size=11, color="FFFFFF")
            elif verdict == "NO MATCH":
                cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                cell.font = Font(bold=True, size=11, color="FFFFFF")
            elif verdict in ["ERROR", "CANNOT_VERIFY"]:
                cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
                cell.font = Font(bold=True, size=11)
        
        # Adjust column width
        ws.column_dimensions['E'].width = 15
        
        print(f"      ✅ Added {len(results)} verdicts")
    
    # Save updated Excel
    print(f"\n💾 Saving: {output_file.name}")
    wb.save(output_file)
    wb.close()
    
    print("\n" + "="*70)
    print("✅ SUCCESS!")
    print("="*70)
    print(f"Updated Excel: {output_file}")
    print()
    print("New column structure:")
    print("  1. Reference A")
    print("  2. Description A")
    print("  3. Reference B")
    print("  4. Description B")
    print("  5. VERDICT (🟢 MATCH / 🔴 NO MATCH)")
    print("  6. Match Confidence %")
    print("  7. Brief Reasoning")
    print()


def extract_from_log():
    """Alternative: Extract verdicts from log file"""
    
    print("\n" + "="*70)
    print("ALTERNATIVE: EXTRACT FROM LOG FILE")
    print("="*70)
    print()
    
    log_file = Path('autonomous_verification.log')
    
    if not log_file.exists():
        print("❌ Log file not found either!")
        print(f"   Looking for: {log_file}")
        print()
        print("Cannot recover verdict data.")
        print("You'll need to re-run (but we can make it much faster!)")
        return
    
    print(f"📂 Reading log: {log_file.name}")
    
    # Parse log for verdicts
    verdicts = {}
    
    with open(log_file, 'r', encoding='utf-8') as f:
        current_pair = None
        
        for line in f:
            # Look for comparison lines
            if "COMPARING:" in line:
                # Reset for new comparison
                current_pair = None
            
            # Extract document pair
            if "<->" in line and current_pair is None:
                parts = line.split("<->")
                if len(parts) == 2:
                    ref_a = parts[0].strip().split()[-1]
                    ref_b = parts[1].strip().split()[0]
                    current_pair = (ref_a, ref_b)
            
            # Extract verdict
            if current_pair and ("MATCH |" in line or "NO MATCH |" in line):
                if "NO MATCH" in line:
                    verdict = "NO MATCH"
                else:
                    verdict = "MATCH"
                
                # Extract confidence
                if "Confidence:" in line:
                    conf_part = line.split("Confidence:")[1].split("%")[0].strip()
                    try:
                        confidence = int(conf_part)
                    except:
                        confidence = 0
                    
                    verdicts[current_pair] = {
                        'verdict': verdict,
                        'confidence': confidence
                    }
    
    if not verdicts:
        print("❌ Could not extract verdicts from log")
        return
    
    print(f"✅ Extracted {len(verdicts)} verdicts from log")
    print()
    
    # Now update Excel
    excel_file = Path(r"C:\Users\JemAndrew\OneDrive - Velitor\Claude\litigation_analysis\cases\lismore_v_ph\analysis\Autonomous_Match_Results.xlsx")
    output_file = Path(r"C:\Users\JemAndrew\OneDrive - Velitor\Claude\litigation_analysis\cases\lismore_v_ph\analysis\Autonomous_Match_Results_WITH_VERDICT.xlsx")
    
    print(f"📊 Loading Excel: {excel_file.name}")
    wb = openpyxl.load_workbook(excel_file)
    
    for sheet_name in wb.sheetnames:
        if sheet_name == "Sheet":
            continue
        
        ws = wb[sheet_name]
        print(f"\n   Processing: {sheet_name}")
        
        # Insert verdict column
        ws.insert_cols(5)
        ws.cell(row=1, column=5).value = "VERDICT"
        ws.cell(row=1, column=5).font = Font(bold=True, size=11, color="FFFFFF")
        ws.cell(row=1, column=5).fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Fill verdicts
        added = 0
        for row_idx in range(2, ws.max_row + 1):
            ref_a = str(ws.cell(row=row_idx, column=1).value).strip()
            ref_b = str(ws.cell(row=row_idx, column=3).value).strip()
            
            pair = (ref_a, ref_b)
            
            if pair in verdicts:
                verdict = verdicts[pair]['verdict']
                
                cell = ws.cell(row=row_idx, column=5)
                cell.value = verdict
                cell.alignment = Alignment(horizontal='center')
                cell.font = Font(bold=True, size=11)
                
                if verdict == "MATCH":
                    cell.fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
                    cell.font = Font(bold=True, size=11, color="FFFFFF")
                else:
                    cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                    cell.font = Font(bold=True, size=11, color="FFFFFF")
                
                added += 1
        
        ws.column_dimensions['E'].width = 15
        print(f"      ✅ Added {added} verdicts")
    
    print(f"\n💾 Saving: {output_file.name}")
    wb.save(output_file)
    wb.close()
    
    print("\n" + "="*70)
    print("✅ SUCCESS!")
    print("="*70)
    print(f"Updated Excel: {output_file}")


if __name__ == "__main__":
    add_verdict_column()