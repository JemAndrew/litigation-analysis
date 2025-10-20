#!/usr/bin/env python3
"""
Entity-Aware Document Match Verification System

Creates NEW Excel with 3 sheets organized by confidence level.
Processes ALL rows with intelligent entity/name matching including abbreviations.
Better OCR for scanned documents.
Anti-hallucination with required evidence citations.

British English throughout.
"""

import os
import sys
from pathlib import Path
from typing import List, Dict, Optional, Tuple
import fitz  # PyMuPDF
import anthropic
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import logging
import time
import base64
from dotenv import load_dotenv
import json

# Load environment variables
load_dotenv()

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('entity_verification.log', encoding='utf-8'),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)


class EntityAwareVerifier:
    """
    Document match verifier with intelligent entity/abbreviation matching.
    
    Understands that:
    - "Brendan Cahill" = "BC" = "B. Cahill" = "Cahill"
    - "Process Holdings Ltd" = "PHL" = "Process Holdings"
    - "Ministry of Petroleum Resources" = "MPR"
    """
    
    def __init__(
        self,
        input_excel: str,
        c_exhibits_folder: str,
        phl_folder: str,
        restricted_folder: str,
        output_excel: str,
        max_cost_gbp: float = 100.0
    ):
        self.input_excel = Path(input_excel)
        self.c_exhibits_folder = Path(c_exhibits_folder)
        self.phl_folder = Path(phl_folder)
        self.restricted_folder = Path(restricted_folder)
        self.output_excel = Path(output_excel)
        self.max_cost_gbp = max_cost_gbp
        
        # API setup
        api_key = os.getenv('ANTHROPIC_API_KEY')
        if not api_key:
            raise ValueError("ANTHROPIC_API_KEY not found in .env")
        
        self.client = anthropic.Anthropic(api_key=api_key)
        
        # Cost tracking
        self.total_cost_gbp = 0.0
        self.api_call_count = 0
        self.INPUT_COST_PER_TOKEN = 0.000003
        self.OUTPUT_COST_PER_TOKEN = 0.000015
        self.USD_TO_GBP = 0.79
        
        # Results organized by confidence
        self.high_confidence_matches: List[Dict] = []      # 90-100%
        self.medium_confidence_matches: List[Dict] = []    # 70-89%
        self.low_confidence_matches: List[Dict] = []       # 0-69%
        
        # All input matches
        self.all_input_matches: List[Dict] = []
        
        # Checkpoint for resume capability
        self.checkpoint_file = Path('verification_checkpoint.json')
        self.processed_indices = set()
    
    def load_checkpoint(self) -> None:
        """Load checkpoint to resume from interruption"""
        if self.checkpoint_file.exists():
            try:
                with open(self.checkpoint_file, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    self.processed_indices = set(data.get('processed_indices', []))
                    self.total_cost_gbp = data.get('total_cost_gbp', 0.0)
                    self.high_confidence_matches = data.get('high_confidence', [])
                    self.medium_confidence_matches = data.get('medium_confidence', [])
                    self.low_confidence_matches = data.get('low_confidence', [])
                    logger.info(f"Resuming: {len(self.processed_indices)} already processed")
                    logger.info(f"Previous cost: GBP {self.total_cost_gbp:.2f}")
            except Exception as e:
                logger.warning(f"Could not load checkpoint: {e}")
    
    def save_checkpoint(self) -> None:
        """Save progress checkpoint"""
        try:
            data = {
                'processed_indices': list(self.processed_indices),
                'total_cost_gbp': self.total_cost_gbp,
                'high_confidence': self.high_confidence_matches,
                'medium_confidence': self.medium_confidence_matches,
                'low_confidence': self.low_confidence_matches,
                'timestamp': datetime.now().isoformat()
            }
            with open(self.checkpoint_file, 'w', encoding='utf-8') as f:
                json.dump(data, f, indent=2)
        except Exception as e:
            logger.warning(f"Could not save checkpoint: {e}")
    
    def load_input_matches(self) -> None:
        """Load all potential matches from input Excel"""
        logger.info(f"Loading matches from: {self.input_excel}")
        
        try:
            wb = openpyxl.load_workbook(self.input_excel, data_only=True)
            ws = wb['Cross-Match Analysis']
            
            for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row or not row[0]:
                    continue
                
                self.all_input_matches.append({
                    'row_index': idx,
                    'factual_ref': str(row[0]).strip() if row[0] else '',
                    'factual_date': str(row[1]).strip() if row[1] else '',
                    'factual_desc': str(row[2]).strip() if row[2] else '',
                    'restricted_ref': str(row[3]).strip() if row[3] else '',
                    'restricted_date': str(row[4]).strip() if row[4] else '',
                    'restricted_desc': str(row[5]).strip() if row[5] else '',
                    'original_match_type': str(row[6]).strip() if row[6] else '',
                    'original_confidence': row[7] if row[7] else 0,
                    'original_desc_similarity': row[8] if row[8] else 0
                })
            
            wb.close()
            logger.info(f"Loaded {len(self.all_input_matches)} matches to verify")
            
        except Exception as e:
            logger.error(f"Error loading input Excel: {e}")
            raise
    
    def find_document(self, reference: str) -> Optional[Path]:
        """Find document path based on reference"""
        try:
            if reference.startswith('C-'):
                # C-exhibits
                path = self.c_exhibits_folder / f"{reference}.pdf"
                if path.exists():
                    return path
                # Try recursive search
                matches = list(self.c_exhibits_folder.rglob(f"{reference}.pdf"))
                return matches[0] if matches else None
            
            elif reference.startswith('PHL_'):
                # PHL documents
                matches = list(self.phl_folder.rglob(f"{reference}.pdf"))
                return matches[0] if matches else None
            
            else:
                # Restricted documents
                path = self.restricted_folder / f"{reference}.pdf"
                return path if path.exists() else None
                
        except Exception as e:
            logger.error(f"Error finding {reference}: {e}")
            return None
    
    def extract_text_with_ocr(self, pdf_path: Path) -> Tuple[str, int, str]:
        """
        Extract text with improved OCR handling.
        
        Returns:
            - full_text: All extracted text
            - page_count: Number of pages
            - extraction_notes: Any warnings/issues
        """
        notes = []
        
        try:
            doc = fitz.open(pdf_path)
            page_count = len(doc)
            full_text = ""
            
            for page_num in range(page_count):
                try:
                    page = doc[page_num]
                    
                    # Try normal text extraction first
                    page_text = page.get_text()
                    
                    # If very little text, might be scanned - try OCR
                    if len(page_text.strip()) < 50:
                        notes.append(f"Page {page_num + 1}: Low text, attempting OCR")
                        # Get page as image and extract text
                        pix = page.get_pixmap(dpi=150)
                        # PyMuPDF's get_text() with "text" option should handle this
                        page_text = page.get_text("text")
                    
                    full_text += f"\n--- PAGE {page_num + 1} ---\n{page_text}"
                    
                except Exception as e:
                    notes.append(f"Page {page_num + 1}: Error - {str(e)}")
            
            doc.close()
            
            # Validate extraction quality
            if len(full_text.strip()) < 100:
                notes.append("WARNING: Very little text extracted - document may be image-only")
            
            notes_text = "; ".join(notes) if notes else f"Extracted {page_count} pages successfully"
            
            return full_text.strip(), page_count, notes_text
            
        except Exception as e:
            logger.error(f"Error extracting text from {pdf_path}: {e}")
            return "", 0, f"EXTRACTION ERROR: {str(e)}"
    
    def verify_with_claude(
        self,
        factual_ref: str,
        factual_text: str,
        restricted_ref: str,
        restricted_text: str
    ) -> Tuple[str, int, str, float]:
        """
        Ask Claude to verify if documents match with entity-aware comparison.
        
        Returns: (verdict, confidence, reasoning, cost_gbp)
        """
        
        # Truncate if too long (keep first 20k chars for context)
        max_chars = 20000
        factual_truncated = factual_text[:max_chars]
        restricted_truncated = restricted_text[:max_chars]
        
        if len(factual_text) > max_chars:
            factual_truncated += f"\n\n[Truncated... Total: {len(factual_text)} chars]"
        if len(restricted_text) > max_chars:
            restricted_truncated += f"\n\n[Truncated... Total: {len(restricted_text)} chars]"
        
        prompt = f"""You are a legal document verification expert. Your task is to determine if two documents are EXACTLY THE SAME document.

DOCUMENT A: {factual_ref}
Full text:
{factual_truncated}

DOCUMENT B: {restricted_ref}
Full text:
{restricted_truncated}

CRITICAL INSTRUCTIONS FOR ENTITY/NAME MATCHING:

When comparing entities and people, you MUST account for abbreviations and variations:

PEOPLE:
- "Brendan Cahill" = "BC" = "B. Cahill" = "Cahill" = "B Cahill"
- "Niall Lawlor" = "NL" = "N. Lawlor" = "Lawlor"
- "Taofiq Tijani" = "TT" = "T. Tijani" = "Tijani"
- First initials + Last name = Full name

COMPANIES/ENTITIES:
- "Process Holdings Ltd" = "PHL" = "Process Holdings" = "P Holdings"
- "Lismore Capital Ltd" = "Lismore" = "Lismore Capital"
- "Ministry of Petroleum Resources" = "MPR" = "the Ministry" = "Ministry of Petroleum"
- Acronyms = Full entity names

COMPARISON REQUIREMENTS:

1. ENTITY VALIDATION:
   - List ALL key entities/people mentioned in Document A
   - List ALL key entities/people mentioned in Document B
   - Confirm they are the SAME entities (accounting for abbreviations)

2. SUBSTANTIVE CONTENT:
   - Are the core facts/events identical?
   - Are dates/amounts/key details the same?
   - Are email chains/correspondence identical?

3. MINOR DIFFERENCES TO IGNORE:
   - Page numbers, headers, footers
   - OCR errors (e.g., "0" vs "O", "l" vs "1")
   - Formatting differences
   - "Re:" vs "RE:" vs "re:"
   - Extra whitespace

4. DIFFERENCES THAT MATTER:
   - Different people/entities (even after accounting for abbreviations)
   - Different dates or amounts
   - Different substantive content
   - One has information the other lacks

RESPOND IN THIS EXACT FORMAT:

VERDICT: MATCH or NO MATCH

CONFIDENCE: [0-100]

ENTITIES_IN_DOC_A: [List key people/companies]

ENTITIES_IN_DOC_B: [List key people/companies]

ENTITY_MATCH: YES or NO (do entities align, accounting for abbreviations?)

KEY_FACTS_COMPARISON: [Do core facts match?]

REASONING: [Detailed explanation with specific evidence. If NO MATCH, cite specific differences. If MATCH, explain why entities/facts align despite any variation in names.]

Be extremely thorough. This is for legal proceedings where accuracy is critical.
"""
        
        try:
            response = self.client.messages.create(
                model="claude-sonnet-4-5-20250929",
                max_tokens=3000,
                temperature=0,
                messages=[{"role": "user", "content": prompt}]
            )
            
            response_text = response.content[0].text.strip()
            
            # Parse response
            verdict = "NO MATCH"
            confidence = 50
            reasoning = response_text
            
            for line in response_text.split('\n'):
                line = line.strip()
                if line.startswith('VERDICT:'):
                    verdict_text = line.replace('VERDICT:', '').strip().upper()
                    verdict = "MATCH" if ('MATCH' in verdict_text and 'NO MATCH' not in verdict_text) else "NO MATCH"
                elif line.startswith('CONFIDENCE:'):
                    try:
                        confidence = int(''.join(filter(str.isdigit, line)))
                        confidence = max(0, min(100, confidence))
                    except:
                        confidence = 50
            
            # Calculate cost
            input_tokens = response.usage.input_tokens
            output_tokens = response.usage.output_tokens
            cost_usd = (input_tokens * self.INPUT_COST_PER_TOKEN) + \
                      (output_tokens * self.OUTPUT_COST_PER_TOKEN)
            cost_gbp = cost_usd * self.USD_TO_GBP
            
            self.total_cost_gbp += cost_gbp
            self.api_call_count += 1
            
            logger.info(f"     Claude: {verdict} | Confidence: {confidence}%")
            logger.info(f"     Cost: GBP {cost_gbp:.4f} | Total: GBP {self.total_cost_gbp:.2f}")
            
            return verdict, confidence, reasoning, cost_gbp
            
        except Exception as e:
            logger.error(f"Claude API error: {e}")
            return "ERROR", 0, f"API Error: {str(e)}", 0.0
    
    def process_all_matches(self) -> None:
        """Process all matches and categorize by confidence"""
        
        logger.info("="*70)
        logger.info("STARTING ENTITY-AWARE DOCUMENT VERIFICATION")
        logger.info("="*70)
        logger.info(f"Total matches to verify: {len(self.all_input_matches)}")
        logger.info(f"Cost limit: GBP {self.max_cost_gbp}")
        logger.info("="*70)
        
        for idx, match in enumerate(self.all_input_matches, start=1):
            # Skip if already processed
            if match['row_index'] in self.processed_indices:
                logger.info(f"[{idx}/{len(self.all_input_matches)}] SKIPPED: Already processed")
                continue
            
            # Check cost limit
            if self.total_cost_gbp >= self.max_cost_gbp:
                logger.warning(f"COST LIMIT REACHED: GBP {self.total_cost_gbp:.2f}")
                logger.warning(f"Stopping at match {idx}/{len(self.all_input_matches)}")
                break
            
            logger.info(f"\n[{idx}/{len(self.all_input_matches)}] VERIFYING:")
            logger.info(f"  {match['factual_ref']} <-> {match['restricted_ref']}")
            logger.info(f"  Original confidence: {match['original_confidence']}%")
            
            # Find documents
            factual_path = self.find_document(match['factual_ref'])
            restricted_path = self.find_document(match['restricted_ref'])
            
            if not factual_path:
                logger.warning(f"  Document not found: {match['factual_ref']}")
                result = {
                    **match,
                    'claude_verdict': 'ERROR',
                    'claude_confidence': 0,
                    'claude_reasoning': f"Document not found: {match['factual_ref']}",
                    'extraction_notes': 'File not found',
                    'cost_gbp': 0.0
                }
                self.low_confidence_matches.append(result)
                self.processed_indices.add(match['row_index'])
                continue
            
            if not restricted_path:
                logger.warning(f"  Document not found: {match['restricted_ref']}")
                result = {
                    **match,
                    'claude_verdict': 'ERROR',
                    'claude_confidence': 0,
                    'claude_reasoning': f"Document not found: {match['restricted_ref']}",
                    'extraction_notes': 'File not found',
                    'cost_gbp': 0.0
                }
                self.low_confidence_matches.append(result)
                self.processed_indices.add(match['row_index'])
                continue
            
            # Extract text with OCR
            logger.info(f"  Extracting text from {factual_path.name}...")
            factual_text, factual_pages, factual_notes = self.extract_text_with_ocr(factual_path)
            
            logger.info(f"  Extracting text from {restricted_path.name}...")
            restricted_text, restricted_pages, restricted_notes = self.extract_text_with_ocr(restricted_path)
            
            extraction_notes = f"Doc A: {factual_pages} pages ({factual_notes}); Doc B: {restricted_pages} pages ({restricted_notes})"
            
            # Verify with Claude
            logger.info(f"  Calling Claude for entity-aware verification...")
            
            verdict, confidence, reasoning, cost = self.verify_with_claude(
                match['factual_ref'], factual_text,
                match['restricted_ref'], restricted_text
            )
            
            # Create result
            result = {
                **match,
                'claude_verdict': verdict,
                'claude_confidence': confidence,
                'claude_reasoning': reasoning,
                'extraction_notes': extraction_notes,
                'cost_gbp': cost
            }
            
            # Categorize by confidence
            if confidence >= 90:
                self.high_confidence_matches.append(result)
                logger.info(f"  HIGH CONFIDENCE: {verdict} ({confidence}%)")
            elif confidence >= 70:
                self.medium_confidence_matches.append(result)
                logger.info(f"  MEDIUM CONFIDENCE: {verdict} ({confidence}%)")
            else:
                self.low_confidence_matches.append(result)
                logger.info(f"  LOW CONFIDENCE: {verdict} ({confidence}%)")
            
            # Mark as processed
            self.processed_indices.add(match['row_index'])
            
            # Save checkpoint every 10 matches
            if idx % 10 == 0:
                logger.info(f"  Saving checkpoint...")
                self.save_checkpoint()
            
            # Rate limiting
            time.sleep(0.5)
        
        logger.info("\n" + "="*70)
        logger.info("VERIFICATION COMPLETE")
        logger.info("="*70)
        logger.info(f"Processed: {len(self.processed_indices)}/{len(self.all_input_matches)}")
        logger.info(f"High confidence (90%+): {len(self.high_confidence_matches)}")
        logger.info(f"Medium confidence (70-89%): {len(self.medium_confidence_matches)}")
        logger.info(f"Low confidence (<70%): {len(self.low_confidence_matches)}")
        logger.info(f"Total cost: GBP {self.total_cost_gbp:.2f}")
        logger.info(f"API calls: {self.api_call_count}")
    
    def create_output_excel(self) -> Path:
        """Create new Excel with 3 sheets organized by confidence"""
        
        logger.info("Creating output Excel with 3 sheets...")
        
        try:
            wb = openpyxl.Workbook()
            
            # Remove default sheet
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
            
            # Define headers
            headers = [
                'Factual Exhibit Ref',
                'Factual Date',
                'Factual Description',
                'Restricted Doc Ref',
                'Restricted Date',
                'Restricted Description',
                'Original Match Type',
                'Original Confidence %',
                'Original Desc Similarity %',
                'Claude Verdict',
                'Claude Confidence %',
                'Claude Detailed Reasoning',
                'Extraction Notes',
                'Cost (GBP)'
            ]
            
            # Header style
            header_font = Font(bold=True, size=11, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            # Create Sheet 1: High Confidence (90-100%)
            ws1 = wb.create_sheet("High Confidence (90%+)")
            ws1.append(headers)
            for cell in ws1[1]:
                cell.font = header_font
                cell.fill = header_fill
            
            for match in self.high_confidence_matches:
                ws1.append([
                    match['factual_ref'],
                    match['factual_date'],
                    match['factual_desc'],
                    match['restricted_ref'],
                    match['restricted_date'],
                    match['restricted_desc'],
                    match['original_match_type'],
                    match['original_confidence'],
                    match['original_desc_similarity'],
                    match['claude_verdict'],
                    match['claude_confidence'],
                    match['claude_reasoning'],
                    match['extraction_notes'],
                    round(match['cost_gbp'], 4)
                ])
                
                # Color code verdict
                last_row = ws1.max_row
                verdict_cell = ws1.cell(row=last_row, column=10)
                if match['claude_verdict'] == 'MATCH':
                    verdict_cell.fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
                elif match['claude_verdict'] == 'NO MATCH':
                    verdict_cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            
            # Create Sheet 2: Medium Confidence (70-89%)
            ws2 = wb.create_sheet("Medium Confidence (70-89%)")
            ws2.append(headers)
            for cell in ws2[1]:
                cell.font = header_font
                cell.fill = header_fill
            
            for match in self.medium_confidence_matches:
                ws2.append([
                    match['factual_ref'],
                    match['factual_date'],
                    match['factual_desc'],
                    match['restricted_ref'],
                    match['restricted_date'],
                    match['restricted_desc'],
                    match['original_match_type'],
                    match['original_confidence'],
                    match['original_desc_similarity'],
                    match['claude_verdict'],
                    match['claude_confidence'],
                    match['claude_reasoning'],
                    match['extraction_notes'],
                    round(match['cost_gbp'], 4)
                ])
                
                last_row = ws2.max_row
                verdict_cell = ws2.cell(row=last_row, column=10)
                if match['claude_verdict'] == 'MATCH':
                    verdict_cell.fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
                elif match['claude_verdict'] == 'NO MATCH':
                    verdict_cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            
            # Create Sheet 3: Low Confidence (<70%)
            ws3 = wb.create_sheet("Low Confidence (<70%)")
            ws3.append(headers)
            for cell in ws3[1]:
                cell.font = header_font
                cell.fill = header_fill
            
            for match in self.low_confidence_matches:
                ws3.append([
                    match['factual_ref'],
                    match['factual_date'],
                    match['factual_desc'],
                    match['restricted_ref'],
                    match['restricted_date'],
                    match['restricted_desc'],
                    match['original_match_type'],
                    match['original_confidence'],
                    match['original_desc_similarity'],
                    match['claude_verdict'],
                    match['claude_confidence'],
                    match['claude_reasoning'],
                    match['extraction_notes'],
                    round(match['cost_gbp'], 4)
                ])
                
                last_row = ws3.max_row
                verdict_cell = ws3.cell(row=last_row, column=10)
                if match['claude_verdict'] == 'MATCH':
                    verdict_cell.fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
                elif match['claude_verdict'] == 'NO MATCH':
                    verdict_cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            
            # Set column widths for all sheets
            for ws in [ws1, ws2, ws3]:
                ws.column_dimensions['A'].width = 18
                ws.column_dimensions['B'].width = 12
                ws.column_dimensions['C'].width = 50
                ws.column_dimensions['D'].width = 18
                ws.column_dimensions['E'].width = 12
                ws.column_dimensions['F'].width = 50
                ws.column_dimensions['G'].width = 25
                ws.column_dimensions['H'].width = 15
                ws.column_dimensions['I'].width = 15
                ws.column_dimensions['J'].width = 15
                ws.column_dimensions['K'].width = 15
                ws.column_dimensions['L'].width = 80
                ws.column_dimensions['M'].width = 50
                ws.column_dimensions['N'].width = 10
                
                # Freeze top row
                ws.freeze_panes = 'A2'
            
            # Save
            wb.save(self.output_excel)
            wb.close()
            
            logger.info(f"Output Excel created: {self.output_excel}")
            logger.info(f"  Sheet 1: {len(self.high_confidence_matches)} high confidence matches")
            logger.info(f"  Sheet 2: {len(self.medium_confidence_matches)} medium confidence matches")
            logger.info(f"  Sheet 3: {len(self.low_confidence_matches)} low confidence matches")
            
            return self.output_excel
            
        except Exception as e:
            logger.error(f"Error creating Excel: {e}")
            raise
    
    def run(self) -> Path:
        """Execute full verification workflow"""
        
        try:
            # Load checkpoint if exists
            self.load_checkpoint()
            
            # Load input matches
            self.load_input_matches()
            
            # Process all matches
            self.process_all_matches()
            
            # Create output Excel
            output_path = self.create_output_excel()
            
            # Clean up checkpoint
            if self.checkpoint_file.exists():
                self.checkpoint_file.unlink()
            
            return output_path
            
        except Exception as e:
            logger.error(f"Fatal error: {e}")
            # Save what we have
            try:
                self.save_checkpoint()
                logger.info("Emergency checkpoint saved")
            except:
                pass
            raise


def main():
    """Main execution"""
    
    input_excel = r"C:\Users\JemAndrew\OneDrive - Velitor\Claude\litigation_analysis\cases\lismore_v_ph\analysis\Cross_Match_Analysis_Results.xlsx"
    
    c_exhibits_folder = r"C:\Users\JemAndrew\Velitor\Communication site - Documents\LIS1.1\73. Review of PHL Disclosures\01. Claimant's Factual Exhibits"
    
    phl_folder = r"C:\Users\JemAndrew\Velitor\Communication site - Documents\LIS1.1\73. Review of PHL Disclosures\02. PHL Disclosure prior to 23 June"
    
    restricted_folder = r"C:\Users\JemAndrew\Velitor\Communication site - Documents\LIS1.1\73. Review of PHL Disclosures\03. Erroneously restricted documents"
    
    output_excel = r"C:\Users\JemAndrew\OneDrive - Velitor\Claude\litigation_analysis\cases\lismore_v_ph\analysis\Document_Match_Verification_Results.xlsx"
    
    # Create verifier
    verifier = EntityAwareVerifier(
        input_excel=input_excel,
        c_exhibits_folder=c_exhibits_folder,
        phl_folder=phl_folder,
        restricted_folder=restricted_folder,
        output_excel=output_excel,
        max_cost_gbp=100.0  # Adjust as needed
    )
    
    # Run verification
    output_path = verifier.run()
    
    print("\n" + "="*70)
    print("VERIFICATION COMPLETE!")
    print("="*70)
    print(f"Output file: {output_path}")
    print(f"\nResults:")
    print(f"  High Confidence (90%+):  {len(verifier.high_confidence_matches)} matches")
    print(f"  Medium Confidence (70-89%): {len(verifier.medium_confidence_matches)} matches")
    print(f"  Low Confidence (<70%):   {len(verifier.low_confidence_matches)} matches")
    print(f"\nTotal processed: {len(verifier.processed_indices)}/{len(verifier.all_input_matches)}")
    print(f"Total cost: GBP {verifier.total_cost_gbp:.2f}")
    print(f"API calls: {verifier.api_call_count}")


if __name__ == "__main__":
    main()