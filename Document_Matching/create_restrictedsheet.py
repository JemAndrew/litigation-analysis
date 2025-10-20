#!/usr/bin/env python3
"""
Cross-Match Analysis: Factual Exhibits vs Restricted Documents

Identifies potential duplicate documents between:
- Factual Exhibits sheet (C-1 to C-156, PHL_000001 to PHL_000554)
- Restricted Documents with Trial Bundle Refs sheet (Folder 03 documents)

Uses fuzzy matching on descriptions and exact date matching.

British English throughout.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path
from typing import List, Dict, Tuple, Optional
from datetime import datetime
from fuzzywuzzy import fuzz
import logging
from dataclasses import dataclass

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


@dataclass
class Document:
    """Represents a document with metadata"""
    reference: str
    date: Optional[str]
    description: str
    source_sheet: str
    row_number: int


@dataclass
class Match:
    """Represents a match between two documents"""
    factual_ref: str
    factual_date: str
    factual_desc: str
    restricted_ref: str
    restricted_date: str
    restricted_desc: str
    match_type: str
    confidence: int
    description_similarity: int
    date_match: bool


class CrossMatchAnalyser:
    """Analyse and identify matches between Factual Exhibits and Restricted Documents"""
    
    def __init__(self, input_excel: str, output_excel: str):
        self.input_excel = Path(input_excel)
        self.output_excel = Path(output_excel)
        self.factual_exhibits: List[Document] = []
        self.restricted_docs: List[Document] = []
        self.matches: List[Match] = []
    
    def load_factual_exhibits(self) -> None:
        """
        Load Factual Exhibits sheet.
        
        Structure: Reference | Date | Description | Trial Bundle Ref | ...
        """
        logger.info("Loading Factual Exhibits...")
        
        try:
            wb = openpyxl.load_workbook(self.input_excel, data_only=True)
            ws = wb['Factual Exhibits']
            
            # Headers in row 1, data starts row 2
            # Column A = Reference, B = Date, C = Description
            
            for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row[0]:  # Skip empty rows
                    continue
                
                reference = str(row[0]).strip()
                
                # Skip section headers (like "Request for Arbitration (12 May 2021)")
                if not reference.startswith(('C-', 'PHL_')):
                    continue
                
                date_value = row[1] if len(row) > 1 else None
                description = str(row[2]).strip() if len(row) > 2 and row[2] else ''
                
                # Parse date
                parsed_date = self._parse_date(date_value)
                
                self.factual_exhibits.append(Document(
                    reference=reference,
                    date=parsed_date,
                    description=description,
                    source_sheet='Factual Exhibits',
                    row_number=idx
                ))
            
            wb.close()
            
            logger.info(f"Loaded {len(self.factual_exhibits)} factual exhibits")
            
        except Exception as e:
            logger.error(f"Error loading Factual Exhibits: {e}")
            raise
    
    def load_restricted_documents(self) -> None:
        """
        Load Restricted Documents with Trial Bundle Refs sheet.
        
        Structure: Disclosure ID | Description | Date
        """
        logger.info("Loading Restricted Documents...")
        
        try:
            wb = openpyxl.load_workbook(self.input_excel, data_only=True)
            ws = wb['Restricted Documents with Trial']
            
            # Headers in row 1, data starts row 2
            # Column A = Disclosure ID, B = Description, C = Date
            
            for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row[0]:  # Skip empty rows
                    continue
                
                disclosure_id = str(row[0]).strip()
                description = str(row[1]).strip() if len(row) > 1 and row[1] else ''
                date_value = row[2] if len(row) > 2 else None
                
                # Skip "NOT FOUND IN TRIAL BUNDLE" entries
                if description == 'NOT FOUND IN TRIAL BUNDLE':
                    continue
                
                # Parse date
                parsed_date = self._parse_date(date_value)
                
                self.restricted_docs.append(Document(
                    reference=disclosure_id,
                    date=parsed_date,
                    description=description,
                    source_sheet='Restricted Documents',
                    row_number=idx
                ))
            
            wb.close()
            
            logger.info(f"Loaded {len(self.restricted_docs)} restricted documents")
            
        except Exception as e:
            logger.error(f"Error loading Restricted Documents: {e}")
            raise
    
    def _parse_date(self, date_value) -> Optional[str]:
        """
        Parse date value to standardised dd/mm/yyyy format.
        
        Handles: datetime objects, Excel serial numbers, string dates.
        Returns None if date cannot be parsed.
        """
        if not date_value or date_value == '':
            return None
        
        # If already a string in dd/mm/yyyy format
        if isinstance(date_value, str):
            date_str = date_value.strip()
            # Try to parse to validate it's a real date
            for fmt in ['%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d', '%d %B %Y', '%d %b %Y']:
                try:
                    dt = datetime.strptime(date_str, fmt)
                    return dt.strftime('%d/%m/%Y')
                except:
                    continue
            return date_str  # Return as-is if can't parse
        
        # If datetime object
        if isinstance(date_value, datetime):
            return date_value.strftime('%d/%m/%Y')
        
        # If Excel serial number
        if isinstance(date_value, (int, float)):
            try:
                from datetime import timedelta
                if date_value > 59:
                    date_value -= 1  # Excel leap year bug
                base_date = datetime(1899, 12, 31)
                actual_date = base_date + timedelta(days=date_value)
                return actual_date.strftime('%d/%m/%Y')
            except:
                return None
        
        return None
    
    def _dates_match(self, date1: Optional[str], date2: Optional[str]) -> bool:
        """Check if two dates match exactly"""
        if not date1 or not date2:
            return False
        return date1 == date2
    
    def _classify_match(
        self, 
        desc_similarity: int, 
        date_match: bool
    ) -> Tuple[str, int]:
        """
        Classify match type based on description similarity and date matching.
        
        Returns: (match_type, confidence_percentage)
        """
        # EXACT MATCH: Date matches + 95%+ description similarity
        if date_match and desc_similarity >= 95:
            return ("EXACT MATCH", 100)
        
        # HIGHLY LIKELY MATCH: Date matches + 80-94% description similarity
        if date_match and desc_similarity >= 80:
            return ("HIGHLY LIKELY MATCH", 90)
        
        # LIKELY MATCH: 85%+ description similarity (regardless of date)
        if desc_similarity >= 85:
            return ("LIKELY MATCH", 75)
        
        # PROBABLE MATCH: 70-84% description similarity
        if desc_similarity >= 70:
            return ("PROBABLE MATCH", 60)
        
        # DATE ONLY MATCH: Date matches but description <70% similar
        if date_match and desc_similarity < 70:
            return ("DATE ONLY MATCH - REVIEW BY HUMAN", 40)
        
        # NO MATCH
        return ("NO MATCH", 0)
    
    def find_matches(self, min_similarity: int = 70) -> None:
        """
        Find all matches between Factual Exhibits and Restricted Documents.
        
        Args:
            min_similarity: Minimum description similarity threshold (0-100)
        """
        logger.info(f"Finding matches (min similarity: {min_similarity}%)...")
        
        total_comparisons = len(self.factual_exhibits) * len(self.restricted_docs)
        logger.info(f"Total comparisons to make: {total_comparisons:,}")
        
        comparison_count = 0
        
        for factual_doc in self.factual_exhibits:
            for restricted_doc in self.restricted_docs:
                comparison_count += 1
                
                # Progress logging every 10,000 comparisons
                if comparison_count % 10000 == 0:
                    logger.info(f"  Progress: {comparison_count:,}/{total_comparisons:,} ({comparison_count/total_comparisons*100:.1f}%)")
                
                # Calculate description similarity using token set ratio
                # (better for legal documents with varying word order)
                desc_similarity = fuzz.token_set_ratio(
                    factual_doc.description.lower(),
                    restricted_doc.description.lower()
                )
                
                # Check date match
                date_match = self._dates_match(factual_doc.date, restricted_doc.date)
                
                # Classify match
                match_type, confidence = self._classify_match(desc_similarity, date_match)
                
                # Only record if above minimum similarity threshold
                if desc_similarity >= min_similarity or date_match:
                    self.matches.append(Match(
                        factual_ref=factual_doc.reference,
                        factual_date=factual_doc.date or '',
                        factual_desc=factual_doc.description,
                        restricted_ref=restricted_doc.reference,
                        restricted_date=restricted_doc.date or '',
                        restricted_desc=restricted_doc.description,
                        match_type=match_type,
                        confidence=confidence,
                        description_similarity=desc_similarity,
                        date_match=date_match
                    ))
        
        logger.info(f"Found {len(self.matches)} potential matches")
        
        # Sort matches by confidence (highest first)
        self.matches.sort(key=lambda m: (m.confidence, m.description_similarity), reverse=True)
        
        # Log match type breakdown
        match_counts = {}
        for match in self.matches:
            match_counts[match.match_type] = match_counts.get(match.match_type, 0) + 1
        
        logger.info("Match type breakdown:")
        for match_type, count in sorted(match_counts.items(), key=lambda x: x[1], reverse=True):
            logger.info(f"  {match_type}: {count}")
    
    def create_output_excel(self) -> Path:
        """
        Create output Excel with match results.
        
        Columns:
        - Factual Exhibit Ref
        - Factual Exhibit Date
        - Factual Exhibit Description
        - Restricted Doc Ref
        - Restricted Doc Date
        - Restricted Doc Description
        - Match Type
        - Confidence %
        - Description Similarity %
        - Date Match (Yes/No)
        """
        logger.info("Creating output Excel file...")
        
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Cross-Match Analysis"
            
            # Define styles
            header_font = Font(bold=True, size=11, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            # Match type colors
            exact_fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")  # Green
            highly_likely_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")  # Light green
            likely_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow
            probable_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")  # Orange
            review_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red
            
            # Write headers
            headers = [
                'Factual Exhibit Ref',
                'Factual Exhibit Date',
                'Factual Exhibit Description',
                'Restricted Doc Ref',
                'Restricted Doc Date',
                'Restricted Doc Description',
                'Match Type',
                'Confidence %',
                'Description Similarity %',
                'Date Match'
            ]
            
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Set column widths
            ws.column_dimensions['A'].width = 18
            ws.column_dimensions['B'].width = 12
            ws.column_dimensions['C'].width = 50
            ws.column_dimensions['D'].width = 18
            ws.column_dimensions['E'].width = 12
            ws.column_dimensions['F'].width = 50
            ws.column_dimensions['G'].width = 30
            ws.column_dimensions['H'].width = 12
            ws.column_dimensions['I'].width = 20
            ws.column_dimensions['J'].width = 12
            
            # Write data
            for idx, match in enumerate(self.matches, start=2):
                ws[f'A{idx}'] = match.factual_ref
                ws[f'B{idx}'] = match.factual_date
                ws[f'C{idx}'] = match.factual_desc
                ws[f'D{idx}'] = match.restricted_ref
                ws[f'E{idx}'] = match.restricted_date
                ws[f'F{idx}'] = match.restricted_desc
                ws[f'G{idx}'] = match.match_type
                ws[f'H{idx}'] = match.confidence
                ws[f'I{idx}'] = match.description_similarity
                ws[f'J{idx}'] = 'Yes' if match.date_match else 'No'
                
                # Apply color coding to Match Type column
                match_cell = ws[f'G{idx}']
                if match.match_type == "EXACT MATCH":
                    match_cell.fill = exact_fill
                elif match.match_type == "HIGHLY LIKELY MATCH":
                    match_cell.fill = highly_likely_fill
                elif match.match_type == "LIKELY MATCH":
                    match_cell.fill = likely_fill
                elif match.match_type == "PROBABLE MATCH":
                    match_cell.fill = probable_fill
                elif "REVIEW BY HUMAN" in match.match_type:
                    match_cell.fill = review_fill
            
            # Freeze top row
            ws.freeze_panes = 'A2'
            
            # Save workbook
            wb.save(self.output_excel)
            wb.close()
            
            logger.info(f"✅ Output Excel created: {self.output_excel}")
            logger.info(f"📊 Total matches recorded: {len(self.matches)}")
            
            return self.output_excel
            
        except Exception as e:
            logger.error(f"Error creating output Excel: {e}")
            raise
    
    def run(self) -> Path:
        """Execute the full cross-match analysis workflow"""
        
        logger.info("=" * 60)
        logger.info("CROSS-MATCH ANALYSIS - START")
        logger.info("=" * 60)
        
        try:
            # Step 1: Load Factual Exhibits
            self.load_factual_exhibits()
            
            # Step 2: Load Restricted Documents
            self.load_restricted_documents()
            
            # Step 3: Find matches
            self.find_matches(min_similarity=70)
            
            # Step 4: Create output Excel
            output_path = self.create_output_excel()
            
            logger.info("=" * 60)
            logger.info("CROSS-MATCH ANALYSIS - COMPLETE")
            logger.info("=" * 60)
            
            return output_path
            
        except Exception as e:
            logger.error(f"FATAL ERROR: {e}")
            raise


def main():
    """Main execution"""
    
    input_excel = r"C:\Users\JemAndrew\OneDrive - Velitor\Claude\litigation_analysis\cases\lismore_v_ph\analysis\Claimants_Exhibits_Matches.xlsx"
    
    output_excel = r"C:\Users\JemAndrew\OneDrive - Velitor\Claude\litigation_analysis\cases\lismore_v_ph\analysis\Cross_Match_Analysis_Results.xlsx"
    
    # Create analyser
    analyser = CrossMatchAnalyser(
        input_excel=input_excel,
        output_excel=output_excel
    )
    
    # Run analysis
    output_path = analyser.run()
    
    print("\n" + "=" * 60)
    print("🎉 CROSS-MATCH ANALYSIS COMPLETE!")
    print("=" * 60)
    print(f"Results saved to: {output_path}")
    print("\nOpen the Excel file to review matches:")
    print("  🟢 Green = EXACT MATCH")
    print("  🟢 Light Green = HIGHLY LIKELY MATCH")
    print("  🟡 Yellow = LIKELY MATCH")
    print("  🟠 Orange = PROBABLE MATCH")
    print("  🔴 Red = DATE ONLY MATCH - REVIEW BY HUMAN")


if __name__ == "__main__":
    main()