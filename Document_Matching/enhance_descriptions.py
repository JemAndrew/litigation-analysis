#!/usr/bin/env python3
"""
Update Claimants Exhibits with Trial Bundle Matches
Adds matched references and Claude-generated litigation summaries to existing Excel file
British English throughout
"""

import pandas as pd
import openpyxl
from pathlib import Path
from datetime import datetime
import shutil
import os
from anthropic import Anthropic
from dotenv import load_dotenv


def generate_litigation_summary_with_claude(client: Anthropic, description: str, date: str, 
                                           page_count: int, doc_ref: str) -> str:
    """
    Use Claude API to generate concise litigation summary from Folder 03 description
    
    Args:
        client: Anthropic API client
        description: Document description from Folder 03
        date: Document date
        page_count: Number of pages
        doc_ref: Document reference (e.g., A2_2)
        
    Returns:
        Claude-generated litigation-style summary for trial bundle
    """
    if not description or description == "NO TEXT EXTRACTED":
        return "Document content not extracted - manual review required"
    
    # Build prompt for Claude
    prompt = f"""You are a litigation analyst creating concise trial bundle descriptions.

Document Reference: {doc_ref}
Date: {date}
Pages: {page_count}

Document Content:
{description[:3000]}  

Task: Create a SHORT, professional trial bundle description (max 15 words + page count).

Style Guide - Match these examples:
- "Letter from Kobre & Kim to Mishcon de Reya dated 2020-05-20"
- "Part 18 Request for Further Information"
- "Response to Part 18 Request"
- "High Court arbitration filing"
- "Witness statement filed in High Court"
- "Agreement dated 2016-08-01"
- "Schedule of payments showing financial transactions"

Your description MUST:
- Be concise (max 15 words before page count)
- Identify document type (letter/court filing/agreement/etc)
- Include sender/recipient for letters
- Include date for letters if available
- End with: ({page_count} page{"s" if page_count > 1 else ""})

Output ONLY the description, nothing else."""

    try:
        response = client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=150,
            temperature=0,
            messages=[
                {"role": "user", "content": prompt}
            ]
        )
        
        summary = response.content[0].text.strip()
        
        # Ensure page count is included
        if f"({page_count} page" not in summary:
            summary += f" ({int(page_count)} page{'s' if page_count > 1 else ''})"
        
        return summary
        
    except Exception as e:
        print(f"      ⚠️ API error for {doc_ref}: {e}")
        # Fallback to basic description
        return f"Legal document ({int(page_count)} page{'s' if page_count > 1 else ''})"


def main():
    """Main processing function"""
    
    print("=" * 70)
    print("UPDATE CLAIMANTS EXHIBITS WITH TRIAL BUNDLE MATCHES")
    print("Using Claude API for intelligent summaries")
    print("=" * 70)
    print()
    
    # Load environment variables from root .env file
    root_path = Path(__file__).parent
    env_file = root_path / ".env"
    
    if not env_file.exists():
        print(f"❌ .env file not found at: {env_file}")
        print("   Please ensure .env contains: ANTHROPIC_API_KEY=your_key")
        return
    
    load_dotenv(env_file)
    
    api_key = os.getenv('ANTHROPIC_API_KEY')
    if not api_key:
        print("❌ ANTHROPIC_API_KEY not found in .env file")
        return
    
    print(f"✅ API key loaded: {api_key[:20]}...")
    print()
    
    # Initialize Anthropic client
    client = Anthropic(api_key=api_key)
    
    # File paths
    base_path = Path(r"C:\Users\JemAndrew\OneDrive - Velitor\Claude\litigation_analysis\cases\lismore_v_ph\analysis")
    
    claimants_file = base_path / "Claimants_Exhibits_ENHANCED_20251015_121858.xlsx"
    folder03_file = base_path / "Folder_03_Enhanced_20251016_111218.xlsx"
    bidirectional_file = base_path / "Bidirectional_Matches_Enhanced_20251016_121342.xlsx"
    
    # Check files exist
    print("Checking files...")
    for file in [claimants_file, folder03_file, bidirectional_file]:
        if not file.exists():
            print(f"❌ File not found: {file}")
            return
        print(f"   ✅ Found: {file.name}")
    print()
    
    # Create backup
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    backup_file = claimants_file.parent / f"Claimants_Exhibits_ENHANCED_20251015_121858_BACKUP_{timestamp}.xlsx"
    
    print(f"Creating backup...")
    shutil.copy2(claimants_file, backup_file)
    print(f"   ✅ Backup created: {backup_file.name}")
    print()
    
    # Load Folder 03 data for lookups
    print("Loading Folder 03 Enhanced...")
    df_folder03 = pd.read_excel(folder03_file)
    folder03_lookup = {}
    for _, row in df_folder03.iterrows():
        folder03_lookup[row['Reference']] = {
            'Description': row['Description'],
            'Date': row['Date'],
            'Page Count': row['Page Count']
        }
    print(f"   ✅ Loaded {len(folder03_lookup)} documents")
    print()
    
    # Load Bidirectional Matches
    print("Loading Bidirectional Matches...")
    df_smoking = pd.read_excel(bidirectional_file, sheet_name='SMOKING_GUNS')
    df_claude = pd.read_excel(bidirectional_file, sheet_name='CLAUDE_APPROVED')
    df_needs = pd.read_excel(bidirectional_file, sheet_name='NEEDS_REVIEW')
    
    print(f"   🔥 Smoking Guns: {len(df_smoking)}")
    print(f"   ✅ Claude Approved: {len(df_claude)}")
    print(f"   ⚠️  Needs Review: {len(df_needs)}")
    
    # Combine all matches
    all_matches = pd.concat([df_smoking, df_claude, df_needs], ignore_index=True)
    print(f"   📊 Total matches: {len(all_matches)}")
    print()
    
    # Group matches by C- exhibit
    matches_by_exhibit = {}
    for _, row in all_matches.iterrows():
        c_exhibit = row['Source_Reference']
        matched_ref = row['Matched_Reference']
        
        if c_exhibit not in matches_by_exhibit:
            matches_by_exhibit[c_exhibit] = []
        matches_by_exhibit[c_exhibit].append(matched_ref)
    
    print(f"C- exhibits with matches: {len(matches_by_exhibit)}")
    print()
    
    # Load Claimants Excel using openpyxl to preserve formatting
    print("Loading Claimants Exhibits (preserving format)...")
    wb = openpyxl.load_workbook(claimants_file)
    ws = wb['Factual Exhibits']
    print(f"   ✅ Loaded worksheet: Factual Exhibits")
    print()
    
    # Examine the actual structure
    print("Examining file structure...")
    print("   Row 1 (header row):")
    for col_idx in range(1, 8):
        cell_value = ws.cell(row=1, column=col_idx).value
        print(f"      Column {col_idx}: '{cell_value}'")
    print()
    
    # Use fixed column positions based on structure
    ref_col = 1         # Column A: C- exhibit reference
    date_col = 2        # Column B: Date
    desc_col = 3        # Column C: Description
    trial_ref_col = 4   # Column D: Trial Bundle Ref
    trial_desc_col = 5  # Column E: Trial Bundle Description
    when_prod_col = 6   # Column F: When Produced
    notes_col = 7       # Column G: Notes
    
    print("Using column structure:")
    print(f"   Column {ref_col} (A): C- Exhibit Reference")
    print(f"   Column {trial_ref_col} (D): Trial Bundle Ref (to update)")
    print(f"   Column {trial_desc_col} (E): Trial Bundle Description (to update)")
    print()
    
    # Find the first data row (skip headers and section titles)
    print("Finding first data row...")
    first_data_row = 2
    for row_idx in range(2, 20):  # Check first 20 rows
        cell_value = str(ws.cell(row=row_idx, column=ref_col).value).strip()
        if cell_value.startswith('C-'):
            first_data_row = row_idx
            print(f"   First C- exhibit found at row {first_data_row}: {cell_value}")
            break
    print()
    
    # Process each row
    print("Processing rows and generating Claude summaries...")
    print("(This may take a few minutes for 18 matches)")
    print("-" * 70)
    
    updated_count = 0
    multiple_matches_count = 0
    total_cost = 0.0
    
    for row_idx in range(first_data_row, ws.max_row + 1):
        c_exhibit_cell = ws.cell(row=row_idx, column=ref_col)
        c_exhibit = str(c_exhibit_cell.value).strip() if c_exhibit_cell.value else ""
        
        # Only process C- exhibits
        if not c_exhibit or not c_exhibit.startswith('C-'):
            continue
        
        if c_exhibit not in matches_by_exhibit:
            continue
        
        # Get matches for this C- exhibit
        matches = matches_by_exhibit[c_exhibit]
        
        # Get current Trial Bundle Ref value
        trial_ref_cell = ws.cell(row=row_idx, column=trial_ref_col)
        current_ref = str(trial_ref_cell.value).strip() if trial_ref_cell.value and str(trial_ref_cell.value).strip() not in ["None", "nan"] else ""
        
        # Determine what to add
        if len(matches) == 1:
            # Single match
            new_ref = matches[0]
            
            # Add to existing or create new
            if current_ref:
                # Add to existing if not already there
                if new_ref not in current_ref:
                    trial_ref_cell.value = f"{current_ref}, {new_ref}"
            else:
                # Set new value
                trial_ref_cell.value = new_ref
            
            # Generate Claude summary and add description
            if new_ref in folder03_lookup:
                doc = folder03_lookup[new_ref]
                
                print(f"   🤖 Generating summary for {c_exhibit} → {new_ref}...")
                
                summary = generate_litigation_summary_with_claude(
                    client,
                    doc['Description'],
                    doc['Date'],
                    doc['Page Count'],
                    new_ref
                )
                
                # Rough cost estimate (input ~1000 tokens, output ~50 tokens)
                # Claude Sonnet 4: $3/M input, $15/M output
                cost = (1000 * 0.000003) + (50 * 0.000015)
                total_cost += cost
                
                trial_desc_cell = ws.cell(row=row_idx, column=trial_desc_col)
                current_desc = str(trial_desc_cell.value).strip() if trial_desc_cell.value and str(trial_desc_cell.value).strip() not in ["None", "nan"] else ""
                
                if not current_desc:
                    trial_desc_cell.value = summary
                else:
                    # Add to existing description if different
                    if summary not in current_desc:
                        trial_desc_cell.value = f"{current_desc}, {summary}"
                
                print(f"      ✅ Summary: {summary[:80]}...")
            
            updated_count += 1
            
        else:
            # Multiple matches - just list them all
            new_refs = ", ".join(matches)
            
            if current_ref:
                trial_ref_cell.value = f"{current_ref}, {new_refs}"
            else:
                trial_ref_cell.value = new_refs
            
            print(f"   ⚠️  {c_exhibit} → {new_refs} (MULTIPLE - manual review needed)")
            multiple_matches_count += 1
            updated_count += 1
    
    print("-" * 70)
    print()
    
    # Save the workbook
    print("Saving updated file...")
    wb.save(claimants_file)
    print(f"   ✅ Saved: {claimants_file.name}")
    print()
    
    # Summary
    print("=" * 70)
    print("✅ UPDATE COMPLETE!")
    print("=" * 70)
    print()
    print(f"📊 Summary:")
    print(f"   Total C- exhibits updated: {updated_count}")
    print(f"   Single matches: {updated_count - multiple_matches_count}")
    print(f"   Multiple matches (needs review): {multiple_matches_count}")
    print(f"   Estimated API cost: £{total_cost * 1.27:.4f} (${total_cost:.4f})")
    print()
    print(f"📄 Files:")
    print(f"   Updated: {claimants_file.name}")
    print(f"   Backup: {backup_file.name}")
    print()
    print("🔍 Next steps:")
    print("   1. Open the updated Excel file")
    print("   2. Review exhibits with multiple matches")
    print("   3. Verify Claude-generated summaries are accurate")
    print("   4. Remove any incorrect matches if needed")


if __name__ == "__main__":
    main()